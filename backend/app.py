"""
AlumniConnect – Flask REST API
Run: python app.py
Endpoints all prefixed /api/
"""

from flask import Flask, request, jsonify, send_from_directory, send_file, g, make_response
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.exceptions import HTTPException, RequestEntityTooLarge
import smtplib
import json
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formatdate, make_msgid
from urllib import error as urllib_error
from urllib import request as urllib_request
from urllib.parse import urlparse
import pymysql
import pymysql.cursors
from pymysql import err as pymysql_err
import config
import os
import uuid
import signal
import random
from threading import Lock
from datetime import date, datetime, timedelta
from openpyxl import load_workbook
import re
from io import BytesIO
import mimetypes
from cloudinary_utils import (
    is_cloudinary_configured, upload_to_cloudinary, 
    build_cloudinary_url, get_fallback_avatar_url, get_fallback_image_url
)

DEFAULT_UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
UPLOAD_FOLDER = os.getenv('UPLOAD_FOLDER', DEFAULT_UPLOAD_FOLDER)
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
ALLOWED_DOCUMENT_EXTENSIONS = {'png', 'jpg', 'jpeg', 'webp'}
ALLOWED_EXISTING_LIST_EXTENSIONS = {'xlsx', 'xls'}
try:
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
except OSError:
    # Vercel filesystem can be read-only outside /tmp.
    UPLOAD_FOLDER = '/tmp/alumniconnect_uploads'
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app = Flask(__name__)
app.secret_key = config.SECRET_KEY
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024  # 32 MB limit
MAX_IMAGE_UPLOAD_BYTES = int(os.getenv('MAX_IMAGE_UPLOAD_BYTES', str(10 * 1024 * 1024)))
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)


def allowed_file(filename, allowed=None):
    """Return True if the filename has an allowed extension.

    - `allowed` may be a set/list of extensions (without dot) or None to use
      the global `ALLOWED_EXTENSIONS`.
    """
    if not filename or not isinstance(filename, str):
        return False
    parts = filename.rsplit('.', 1)
    if len(parts) < 2:
        return False
    ext = parts[1].lower()
    allowed_set = set(allowed) if allowed else ALLOWED_EXTENSIONS
    return ext in allowed_set

# Avoid browser "Failed to fetch" from CORS misconfiguration.
# If origins is wildcard, do not enable credentials (invalid in CORS).
def _expanded_cors_origins(origins):
    expanded = []
    for origin in origins:
        value = (origin or '').strip().rstrip('/')
        if not value:
            continue

        expanded.append(value)

    return expanded


_allow_all_origins = '*' in config.CORS_ORIGINS
_cors_origins = '*' if _allow_all_origins else _expanded_cors_origins(config.CORS_ORIGINS)
CORS(
    app,
    resources={r"/api/*": {"origins": _cors_origins}, r"/uploads/*": {"origins": _cors_origins}},
    supports_credentials=not _allow_all_origins,
    methods=['GET', 'POST', 'PUT', 'PATCH', 'DELETE', 'OPTIONS'],
    allow_headers=['Content-Type', 'Authorization'],
    expose_headers=['Content-Type', 'Content-Length'],
)

_db_ready = False
_db_init_attempted = False  # Prevent retry loops on failed initialization
_db_ready_lock = Lock()


def _mysql_ssl_kwargs():
    mode = (config.MYSQL_SSL_MODE or '').strip().lower()
    if not mode or mode == 'disable':
        return None

    if mode in {'required', 'verify_ca', 'verify_identity'}:
        ssl_kwargs = {}
        if config.MYSQL_SSL_CA:
            ssl_kwargs['ca'] = config.MYSQL_SSL_CA
        if mode == 'verify_identity':
            ssl_kwargs['check_hostname'] = True
        elif mode == 'verify_ca':
            ssl_kwargs['check_hostname'] = False
        return ssl_kwargs

    return None


def _mysql_connect_kwargs(include_database=False, dict_cursor=False, autocommit=False):
    kwargs = {
        'host': config.MYSQL_HOST,
        'user': config.MYSQL_USER,
        'password': config.MYSQL_PASSWORD,
        'port': config.MYSQL_PORT,
        'charset': 'utf8mb4',
        'connect_timeout': config.MYSQL_CONNECT_TIMEOUT,
        'autocommit': autocommit,
    }
    if config.MYSQL_READ_TIMEOUT and config.MYSQL_READ_TIMEOUT > 0:
        kwargs['read_timeout'] = config.MYSQL_READ_TIMEOUT
    if config.MYSQL_WRITE_TIMEOUT and config.MYSQL_WRITE_TIMEOUT > 0:
        kwargs['write_timeout'] = config.MYSQL_WRITE_TIMEOUT
    if include_database:
        kwargs['database'] = config.MYSQL_DB
    if dict_cursor:
        kwargs['cursorclass'] = pymysql.cursors.DictCursor

    ssl_kwargs = _mysql_ssl_kwargs()
    if ssl_kwargs is not None:
        kwargs['ssl'] = ssl_kwargs

    return kwargs


_TRANSIENT_MYSQL_ERROR_CODES = {
    1047,  # Unknown command (server restart races)
    1152,  # Aborted connection
    1153,  # Packet too large/network disruption
    1205,  # Lock wait timeout
    1213,  # Deadlock found
    2003,  # Can't connect to MySQL server
    2005,  # Unknown MySQL server host
    2006,  # MySQL server has gone away
    2013,  # Lost connection during query
}


def _is_transient_mysql_error(exc):
    if not isinstance(exc, pymysql_err.OperationalError):
        return False
    code = exc.args[0] if exc.args else None
    return code in _TRANSIENT_MYSQL_ERROR_CODES


def _connect_mysql_with_retry(include_database=False, dict_cursor=False, autocommit=False):
    attempts = max(0, int(config.DB_CONNECT_RETRIES)) + 1
    last_error = None

    for attempt in range(1, attempts + 1):
        try:
            return pymysql.connect(
                **_mysql_connect_kwargs(
                    include_database=include_database,
                    dict_cursor=dict_cursor,
                    autocommit=autocommit,
                )
            )
        except pymysql_err.OperationalError as e:
            last_error = e
            if attempt >= attempts or not _is_transient_mysql_error(e):
                raise

            backoff = config.DB_RETRY_BASE_DELAY * (2 ** (attempt - 1))
            sleep_seconds = min(config.DB_RETRY_MAX_DELAY, max(0.1, backoff))
            sleep_seconds += random.uniform(0, min(0.25, sleep_seconds * 0.2))
            code = e.args[0] if e.args else '?'
            print(f"[DB CONNECT] transient error {code}; retry {attempt}/{attempts - 1} in {sleep_seconds:.2f}s")
            time.sleep(sleep_seconds)

    raise last_error


def init_db_tables_from_schema():
    """Ensure DB and tables exist by executing schema.sql (safe with IF NOT EXISTS/INSERT IGNORE)."""
    schema_path = os.path.join(os.path.dirname(__file__), 'schema.sql')
    if not os.path.exists(schema_path):
        return

    with open(schema_path, 'r', encoding='utf-8') as f:
        schema_sql = f.read()

    # Drop comment lines first so statements after comments are not skipped.
    filtered_lines = []
    for line in schema_sql.splitlines():
        stripped = line.strip()
        if stripped.startswith('--'):
            continue
        filtered_lines.append(line)

    # schema.sql contains local CREATE DATABASE/USE lines; those are handled explicitly here.
    statements = []
    filtered_sql = '\n'.join(filtered_lines)
    for chunk in filtered_sql.split(';'):
        stmt = chunk.strip()
        if not stmt:
            continue
        upper = stmt.upper()
        if upper.startswith('CREATE DATABASE') or upper.startswith('USE '):
            continue
        statements.append(stmt)

    # Prefer connecting to an existing DB first (works for managed providers where CREATE DATABASE is restricted).
    db_conn = None
    try:
        db_conn = _connect_mysql_with_retry(include_database=True, autocommit=True)
    except pymysql_err.OperationalError as e:
        code = e.args[0] if e.args else None
        if code != 1049:  # Unknown database
            raise

        # Database does not exist. Try to create it if privileges allow.
        bootstrap_conn = _connect_mysql_with_retry(include_database=False, autocommit=True)
        try:
            cur = bootstrap_conn.cursor()
            cur.execute(f"CREATE DATABASE IF NOT EXISTS `{config.MYSQL_DB}` CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci")
            cur.close()
        finally:
            bootstrap_conn.close()

        db_conn = _connect_mysql_with_retry(include_database=True, autocommit=True)

    try:
        cur = db_conn.cursor()
        for stmt in statements:
            cur.execute(stmt)
        cur.close()
    finally:
        db_conn.close()


def ensure_db_migrations():
    """Apply small safe migrations for already-existing databases."""
    conn = _connect_mysql_with_retry(include_database=True, autocommit=True)
    try:
        cur = conn.cursor()
        
        # Quick check: if alumni table doesn't exist, abort migrations
        # (schema.sql hasn't been applied yet)
        try:
            cur.execute("SELECT 1 FROM alumni LIMIT 1")
        except pymysql_err.ProgrammingError as e:
            if e.args[0] == 1146:  # Table doesn't exist
                print("[DB MIGRATION] Skipped: alumni table not found (schema.sql not yet applied)")
                return
            raise
        
        # Attempt to add current_job_start_date column if missing
        try:
            cur.execute("ALTER TABLE alumni ADD COLUMN current_job_start_date DATE DEFAULT NULL")
        except pymysql_err.OperationalError as e:
            if e.args[0] != 1060:  # Column already exists
                print(f"[DB MIGRATION] Note (non-fatal): {e}")

        # Create optional helper tables 
        create_stmts = [
            """CREATE TABLE IF NOT EXISTS past_job_experiences (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                alumni_id   INT NOT NULL,
                company     VARCHAR(150) DEFAULT NULL,
                designation VARCHAR(150) DEFAULT NULL,
                start_date  DATE DEFAULT NULL,
                end_date    DATE DEFAULT NULL,
                created_at  DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (alumni_id) REFERENCES alumni(id) ON DELETE CASCADE
            )""",
            """CREATE TABLE IF NOT EXISTS email_logs (
                id              INT AUTO_INCREMENT PRIMARY KEY,
                subject         VARCHAR(255) NOT NULL,
                recipient_count INT NOT NULL DEFAULT 0,
                sent_count      INT NOT NULL DEFAULT 0,
                failed_count    INT NOT NULL DEFAULT 0,
                status          VARCHAR(20) NOT NULL,
                error_message   TEXT DEFAULT NULL,
                created_at      DATETIME DEFAULT CURRENT_TIMESTAMP
            )""",
            """CREATE TABLE IF NOT EXISTS existing_lists (
                id              INT AUTO_INCREMENT PRIMARY KEY,
                title           VARCHAR(255) NOT NULL,
                file_name       VARCHAR(255) NOT NULL,
                stored_path     VARCHAR(500) NOT NULL,
                uploaded_by     VARCHAR(100) DEFAULT 'admin',
                created_at      DATETIME DEFAULT CURRENT_TIMESTAMP
            )""",
            """CREATE TABLE IF NOT EXISTS fund_requests (
                id                   INT AUTO_INCREMENT PRIMARY KEY,
                title                VARCHAR(200) NOT NULL,
                purpose              TEXT NOT NULL,
                target_amount        DECIMAL(12,2) NOT NULL,
                payment_option       ENUM('bkash','bank','both') DEFAULT 'both',
                bkash_number         VARCHAR(30) DEFAULT NULL,
                bank_account_name    VARCHAR(150) DEFAULT NULL,
                bank_account_number  VARCHAR(80) DEFAULT NULL,
                bank_name            VARCHAR(150) DEFAULT NULL,
                status               ENUM('open','closed') DEFAULT 'open',
                created_by           VARCHAR(100) DEFAULT 'admin',
                created_at           DATETIME DEFAULT CURRENT_TIMESTAMP
            )""",
            """CREATE TABLE IF NOT EXISTS password_reset_otps (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                alumni_id   INT NOT NULL,
                email       VARCHAR(255) NOT NULL,
                user_type   ENUM('student','alumni') NOT NULL,
                otp_hash    VARCHAR(255) NOT NULL,
                expires_at  DATETIME NOT NULL,
                used        TINYINT(1) NOT NULL DEFAULT 0,
                attempts    INT NOT NULL DEFAULT 0,
                created_at  DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (alumni_id) REFERENCES alumni(id) ON DELETE CASCADE,
                INDEX idx_password_reset_lookup (email, user_type, used, expires_at)
            )""",
            """CREATE TABLE IF NOT EXISTS uploaded_files (
                id            INT AUTO_INCREMENT PRIMARY KEY,
                file_key      VARCHAR(255) NOT NULL UNIQUE,
                original_name VARCHAR(255) DEFAULT NULL,
                content_type  VARCHAR(120) DEFAULT NULL,
                file_data     LONGBLOB NOT NULL,
                created_at    DATETIME DEFAULT CURRENT_TIMESTAMP
            )""",
            """CREATE TABLE IF NOT EXISTS alumni_referrals (
                id                      INT AUTO_INCREMENT PRIMARY KEY,
                referred_by_alumni_id   INT NOT NULL,
                referred_name           VARCHAR(120) NOT NULL,
                referred_email          VARCHAR(150) NOT NULL,
                referred_phone          VARCHAR(30) DEFAULT NULL,
                referred_student_id     VARCHAR(30) DEFAULT NULL,
                referred_session        VARCHAR(20) DEFAULT NULL,
                referred_department     VARCHAR(50) DEFAULT 'ICE',
                relation_note           VARCHAR(255) DEFAULT NULL,
                status                  ENUM('pending','approved','rejected') DEFAULT 'pending',
                admin_note              VARCHAR(255) DEFAULT NULL,
                reviewed_by             VARCHAR(60) DEFAULT NULL,
                reviewed_at             DATETIME DEFAULT NULL,
                created_at              DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (referred_by_alumni_id) REFERENCES alumni(id) ON DELETE CASCADE,
                INDEX idx_referral_pending (status, created_at),
                INDEX idx_referral_email (referred_email),
                INDEX idx_referral_student_id (referred_student_id)
            )""",
        ]
        
        for stmt in create_stmts:
            try:
                cur.execute(stmt)
            except Exception as e:
                print(f"[DB MIGRATION] CREATE TABLE warning (non-fatal): {e}")

        # Attempt to add fund_transactions columns if missing
        fund_tx_columns = [
            'request_id', 'alumni_id', 'payment_method', 'payment_reference', 
            'created_by_role', 'status'
        ]
        for col in fund_tx_columns:
            try:
                cur.execute(f"ALTER TABLE fund_transactions ADD COLUMN {col} VARCHAR(100) DEFAULT NULL")
            except pymysql_err.OperationalError as e:
                if e.args[0] != 1060:  # Column already exists is OK
                    print(f"[DB MIGRATION] ALTER TABLE note (non-fatal): {e}")

        # Add is_manually_added to alumni if missing  
        try:
            cur.execute("ALTER TABLE alumni ADD COLUMN is_manually_added TINYINT DEFAULT 0")
        except pymysql_err.OperationalError as e:
            if e.args[0] != 1060:  # Column already exists is OK
                print(f"[DB MIGRATION] Note (non-fatal): {e}")

        try:
            cur.execute("ALTER TABLE alumni ADD COLUMN address VARCHAR(255) DEFAULT NULL")
        except pymysql_err.OperationalError as e:
            if e.args[0] != 1060:
                print(f"[DB MIGRATION] Note (non-fatal): {e}")

        try:
            cur.execute("ALTER TABLE alumni ADD COLUMN higher_study VARCHAR(255) DEFAULT NULL")
        except pymysql_err.OperationalError as e:
            if e.args[0] != 1060:
                print(f"[DB MIGRATION] Note (non-fatal): {e}")

        # Add optional event time column if missing
        try:
            cur.execute("ALTER TABLE events ADD COLUMN event_time TIME DEFAULT NULL")
        except pymysql_err.OperationalError as e:
            if e.args[0] != 1060:  # Column already exists is OK
                print(f"[DB MIGRATION] Note (non-fatal): {e}")

        # Legacy deployments may miss events.created_at; keep sorting-compatible schema.
        try:
            cur.execute("ALTER TABLE events ADD COLUMN created_at DATETIME DEFAULT CURRENT_TIMESTAMP")
        except pymysql_err.OperationalError as e:
            if e.args[0] != 1060:  # Column already exists is OK
                print(f"[DB MIGRATION] Note (non-fatal): {e}")

        cur.close()
    finally:
        conn.close()


if config.AUTO_INIT_DB:
    try:
        init_db_tables_from_schema()
    except Exception as e:
        # Keep startup resilient; API endpoints will return DB errors if connectivity is still wrong.
        print(f"[DB INIT] Skipped due to error: {e}")

try:
    ensure_db_migrations()
except Exception as e:
    print(f"[DB MIGRATION] Skipped due to error: {e}")


def ensure_runtime_db_ready():
    """Best-effort runtime DB bootstrap for managed deploys.

    This helps recover from transient startup ordering issues where DB was not
    reachable during initial boot but becomes reachable later.
    
    CRITICAL: Set _db_init_attempted regardless of success to prevent infinite retries.
    """
    global _db_ready, _db_init_attempted
    if _db_ready or _db_init_attempted:
        return

    with _db_ready_lock:
        if _db_ready or _db_init_attempted:
            return

        _db_init_attempted = True  # Mark attempt BEFORE trying, so failures don't retry
        try:
            if config.AUTO_INIT_DB:
                init_db_tables_from_schema()
            ensure_db_migrations()
            _db_ready = True
        except Exception as e:
            print(f"[DB RUNTIME INIT] ERROR (won't retry): {e}")
            import traceback
            traceback.print_exc()  # Print full stack trace for debugging


def build_upload_url(filename):
    """
    Build a complete URL for uploaded images.
    Supports Cloudinary URLs, local filesystem URLs, and HTTP(S) URLs.
    Returns a fallback placeholder if the image is not found or filename is invalid.
    """
    if not filename:
        return get_fallback_image_url('profile')

    raw_value = str(filename).strip()
    if not raw_value:
        return get_fallback_image_url('profile')

    # If already a full HTTP(S) URL, return as-is
    if raw_value.lower().startswith(('http://', 'https://')):
        return raw_value

    # If it's a Cloudinary public_id (contains /) or starts with cloudinary prefix, build URL
    if 'cloudinary' in raw_value or raw_value.startswith('alumniconnect'):
        try:
            url = build_cloudinary_url(raw_value)
            if url:
                return url
        except Exception:
            pass

    # Fallback: try to build local filesystem URL (for backward compatibility)
    parsed = urlparse(raw_value)
    normalized_key = parsed.path or raw_value
    normalized_key = normalized_key.strip().lstrip('/')
    if normalized_key.startswith('api/uploads/'):
        normalized_key = normalized_key[len('api/uploads/'):]
    elif normalized_key.startswith('uploads/'):
        normalized_key = normalized_key[len('uploads/'):]

    if not normalized_key:
        return get_fallback_image_url('profile')

    # Construct base URL - prioritize PUBLIC_BASE_URL env var
    if config.PUBLIC_BASE_URL:
        base_url = config.PUBLIC_BASE_URL.rstrip('/')
    else:
        # Use request context to build URL
        base_url = request.host_url.rstrip('/')
        # For reverse proxies, check X-Forwarded-Proto header
        if request.headers.get('X-Forwarded-Proto'):
            proto = request.headers.get('X-Forwarded-Proto')
            host = request.headers.get('X-Forwarded-Host') or request.host
            base_url = f"{proto}://{host}"
    
    # Remove /api suffix if present
    if base_url.endswith('/api'):
        base_url = base_url[:-4]
    
    url = f"{base_url}/uploads/{normalized_key}"
    print(f"[UPLOAD URL] Built: {url} from filename: {filename}")
    return url


def save_uploaded_image(file_storage):
    """
    Save uploaded image to Cloudinary (primary) with fallback to local filesystem.
    Returns the image identifier (Cloudinary public_id or local filename).
    """
    if not file_storage or not file_storage.filename:
        return None

    ext = file_storage.filename.rsplit('.', 1)[-1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        return None

    safe_name = secure_filename(file_storage.filename)
    file_bytes = file_storage.read()
    if not file_bytes:
        return None
    if len(file_bytes) > MAX_IMAGE_UPLOAD_BYTES:
        return None

    # PRIMARY: Try to upload to Cloudinary for production reliability
    if is_cloudinary_configured():
        try:
            cloudinary_result = upload_to_cloudinary(
                file_bytes,
                filename=safe_name,
                folder='alumniconnect',
            )
            if not cloudinary_result.get('error'):
                public_id = cloudinary_result.get('public_id')
                if public_id:
                    print(f"[UPLOAD] Cloudinary success: {public_id}")
                    return public_id
            else:
                print(f"[UPLOAD] Cloudinary failed: {cloudinary_result.get('error')}")
        except Exception as e:
            print(f"[UPLOAD] Cloudinary exception: {e}")
    
    # FALLBACK: Save locally for dev/testing and as backup
    file_key = f"{uuid.uuid4().hex}_{safe_name}"
    
    # Try filesystem
    try:
        local_path = os.path.join(UPLOAD_FOLDER, file_key)
        os.makedirs(os.path.dirname(local_path), exist_ok=True)
        with open(local_path, 'wb') as out:
            out.write(file_bytes)
        print(f"[UPLOAD] Local filesystem saved: {file_key}")
    except OSError as e:
        print(f"[UPLOAD] Local filesystem failed (non-fatal): {e}")
        # Non-fatal in serverless environments; Cloudinary is preferred
    
    # Try database as additional backup
    db_saved = False
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS uploaded_files (
                id            INT AUTO_INCREMENT PRIMARY KEY,
                file_key      VARCHAR(255) NOT NULL UNIQUE,
                original_name VARCHAR(255) DEFAULT NULL,
                content_type  VARCHAR(120) DEFAULT NULL,
                file_data     LONGBLOB NOT NULL,
                created_at    DATETIME DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        cur.execute(
            """
            INSERT INTO uploaded_files (file_key, original_name, content_type, file_data)
            VALUES (%s, %s, %s, %s)
            """,
            (
                file_key,
                file_storage.filename,
                (file_storage.mimetype or '').strip() or None,
                file_bytes,
            ),
        )
        conn.commit()
        cur.close()
        db_saved = True
        print(f"[UPLOAD] Database backup saved: {file_key}")
    except Exception as e:
        print(f"[UPLOAD] Database backup failed (non-fatal): {e}")

    return file_key


def fetch_uploaded_file(file_key):
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT file_data, content_type FROM uploaded_files WHERE file_key=%s LIMIT 1",
            (file_key,),
        )
        row = cur.fetchone()
        return row
    except Exception:
        return None
    finally:
        cur.close()


def send_registration_received_email(name, email, user_type):
    role_label = 'student' if user_type == 'student' else 'alumni'
    send_email(
        [email],
        'Registration request received - AlumniConnect',
        (
            f"Hi {name or 'there'},\n\n"
            f"We received your {role_label} registration request. "
            "Your account is now pending admin approval.\n\n"
            "You will get another email once your account is approved."
        ),
        'Your registration is in review.',
        'Visit AlumniConnect'
    )


def send_existing_alumni_activation_email(name, email):
    send_email(
        [email],
        'Welcome to AlumniConnect - Your account is now active',
        (
            f"Hi {name or 'there'},\n\n"
            "Your registration matched our existing alumni record. "
            "Your account is now active and moved to All Alumni.\n\n"
            "Please login to continue."
        ),
        'Your account is active and ready to use.',
        'Login to AlumniConnect'
    )


def send_referral_invitation_email(name, email):
    send_email(
        [email],
        'You are referred for joining AlumniConnect',
        (
            f"Hi {name or 'there'},\n\n"
            "You are referred for joining and connecting with all ICE alumni and students. "
            "Please open AlumniConnect and register your account."
        ),
        'You have a new AlumniConnect referral invitation.',
        'Open AlumniConnect'
    )


def fetch_past_jobs(conn, alumni_id):
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, company, designation, start_date, end_date, created_at
        FROM past_job_experiences
        WHERE alumni_id=%s
        ORDER BY COALESCE(end_date, '9999-12-31') DESC, id DESC
        """,
        (alumni_id,),
    )
    rows = cur.fetchall()
    cur.close()
    for r in rows:
        if r.get('start_date'):
            r['start_date'] = str(r['start_date'])
        if r.get('end_date'):
            r['end_date'] = str(r['end_date'])
        if r.get('created_at'):
            r['created_at'] = str(r['created_at'])
    return rows


def insert_email_log(conn, subject, recipient_count, sent_count, failed_count, status, error_message=None):
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO email_logs (subject, recipient_count, sent_count, failed_count, status, error_message)
        VALUES (%s,%s,%s,%s,%s,%s)
        """,
        (
            (subject or '')[:255],
            int(recipient_count or 0),
            int(sent_count or 0),
            int(failed_count or 0),
            (status or 'failed')[:20],
            (str(error_message)[:1000] if error_message else None),
        ),
    )

    # Keep only latest 10 logs.
    cur.execute(
        """
        DELETE FROM email_logs
        WHERE id NOT IN (
            SELECT id FROM (
                SELECT id
                FROM email_logs
                ORDER BY created_at DESC, id DESC
                LIMIT 10
            ) AS keep_rows
        )
        """
    )
    conn.commit()
    cur.close()


def fetch_email_logs(conn, limit=10):
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, subject, recipient_count, sent_count, failed_count, status, error_message, created_at
        FROM email_logs
        ORDER BY created_at DESC, id DESC
        LIMIT %s
        """,
        (int(limit),),
    )
    rows = cur.fetchall()
    cur.close()
    for row in rows:
        if row.get('created_at'):
            row['created_at'] = str(row['created_at'])
    return rows


def smtp_configured():
    return bool(
        config.SMTP_HOST
        and config.SMTP_PORT
        and config.SMTP_USERNAME
        and config.SMTP_PASSWORD
        and config.SMTP_FROM_EMAIL
    )


def brevo_configured():
    return bool(config.BREVO_API_KEY and config.SMTP_FROM_EMAIL)


def active_mail_provider():
    provider = (config.MAIL_PROVIDER or 'auto').strip().lower()
    if provider in {'smtp', 'brevo'}:
        return provider
    # In auto mode prefer SMTP when available; it is often accepted better by institutional inboxes.
    if smtp_configured():
        return 'smtp'
    if brevo_configured():
        return 'brevo'
    return 'smtp'


def _build_email_html(subject, preheader, message, cta_text='Visit AlumniConnect'):
    escaped_subject = str(subject or '')
    escaped_preheader = str(preheader or '')
    escaped_body = str(message or '').replace('\n', '<br>')
    escaped_cta = str(cta_text or 'Visit AlumniConnect')
    unsubscribe_href = config.MAIL_UNSUBSCRIBE_URL or (f"mailto:{config.MAIL_UNSUBSCRIBE_EMAIL}?subject=unsubscribe" if config.MAIL_UNSUBSCRIBE_EMAIL else '')
    footer_lines = [
        str(config.MAIL_FOOTER_TEXT or ''),
        f"Reply contact: {config.MAIL_REPLY_TO}" if config.MAIL_REPLY_TO else '',
    ]
    footer_html = '<br>'.join([line for line in footer_lines if line])
    base_url = config.PUBLIC_BASE_URL or 'http://localhost:5173'
    return f"""
    <html>
        <body style=\"margin:0;padding:0;background:#f7f2ff;font-family:Inter,Segoe UI,Arial,sans-serif;color:#2d0a50;\">
            <table role=\"presentation\" width=\"100%\" cellspacing=\"0\" cellpadding=\"0\" style=\"padding:24px 10px;\">
                <tr>
                    <td align=\"center\">
                            <tr>
                                <td style="padding:0 26px 24px;color:#7b6898;font-size:12px;line-height:1.7;border-top:1px solid #f0e6fb;">
                                    <div style="padding-top:14px;">{footer_html}</div>
                                    {f'<div style="margin-top:8px;"><a href="{unsubscribe_href}" style="color:#6b3f9a;">Unsubscribe</a></div>' if unsubscribe_href else ''}
                                </td>
                            </tr>
                        <table role=\"presentation\" width=\"640\" cellspacing=\"0\" cellpadding=\"0\" style=\"max-width:640px;background:#ffffff;border-radius:16px;overflow:hidden;border:1px solid #ecdffb;\">
                            <tr>
                                <td style=\"padding:20px 26px;background:linear-gradient(135deg,#4f1c78,#a4508b);color:white;\">
                                    <div style=\"font-size:12px;letter-spacing:1px;opacity:0.8;font-weight:700;\">ALUMNICONNECT</div>
                                    <div style=\"font-size:24px;line-height:1.3;font-weight:700;margin-top:8px;\">{escaped_subject}</div>
                                    <div style=\"font-size:13px;line-height:1.6;opacity:0.9;margin-top:8px;\">{escaped_preheader}</div>
                                </td>
                            </tr>
                            <tr>
                                <td style=\"padding:26px;color:#4a3764;font-size:15px;line-height:1.7;\">{escaped_body}</td>
                            </tr>
                            <tr>
                                <td style=\"padding:0 26px 26px;\">
                                    <a href=\"{base_url}\" style=\"display:inline-block;padding:11px 18px;background:linear-gradient(135deg,#5f2c82,#a4508b);color:#fff;text-decoration:none;border-radius:999px;font-weight:700;font-size:13px;\">{escaped_cta}</a>
                                </td>
                            </tr>
                        </table>
                    </td>
                </tr>
            </table>
        </body>
    </html>
    """


def _normalize_recipient_emails(raw_recipients):
    normalized = []
    seen = set()
    pattern = re.compile(r'^[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}$', re.IGNORECASE)

    for raw in (raw_recipients or []):
        email = str(raw or '').strip().lower()
        if not email or email in seen:
            continue
        if not pattern.match(email):
            continue
        seen.add(email)
        normalized.append(email)

    return normalized


def _normalize_plain_content(subject, preheader, message):
    blocks = [str(subject or '').strip(), str(preheader or '').strip(), str(message or '').strip()]
    filtered = [b for b in blocks if b]
    tail_parts = [str(config.MAIL_FOOTER_TEXT or '').strip()]
    if config.MAIL_REPLY_TO:
        tail_parts.append(f"Reply: {config.MAIL_REPLY_TO}")
    if config.MAIL_UNSUBSCRIBE_EMAIL:
        tail_parts.append(f"Unsubscribe: mailto:{config.MAIL_UNSUBSCRIBE_EMAIL}?subject=unsubscribe")
    tail = "\n".join([p for p in tail_parts if p])
    joined = "\n\n".join(filtered)
    if tail:
        joined = f"{joined}\n\n{tail}" if joined else tail
    return joined + "\n"


def _mail_metadata_headers():
    headers = {
        'Precedence': 'bulk',
        'X-Auto-Response-Suppress': 'All',
        'List-ID': config.MAIL_LIST_ID,
        'Feedback-ID': config.MAIL_FEEDBACK_ID,
    }

    unsubscribe_parts = []
    if config.MAIL_UNSUBSCRIBE_URL:
        unsubscribe_parts.append(f"<{config.MAIL_UNSUBSCRIBE_URL}>")
    if config.MAIL_UNSUBSCRIBE_EMAIL:
        unsubscribe_parts.append(f"<mailto:{config.MAIL_UNSUBSCRIBE_EMAIL}?subject=unsubscribe>")

    if unsubscribe_parts:
        headers['List-Unsubscribe'] = ', '.join(unsubscribe_parts)
        headers['List-Unsubscribe-Post'] = 'List-Unsubscribe=One-Click'

    return headers


def send_email(to_emails, subject, message, preheader='', cta_text='Visit AlumniConnect'):
    recipients = _normalize_recipient_emails(to_emails)
    if not recipients:
        return {'sent': 0, 'failed': 0, 'errors': []}

    html_content = _build_email_html(subject, preheader, message, cta_text)
    plain_content = _normalize_plain_content(subject, preheader, message)
    provider = active_mail_provider()
    forced_domains = {d.strip().lower() for d in (config.MAIL_FORCE_SMTP_DOMAINS or []) if d and d.strip()}

    forced_smtp_recipients = []
    regular_recipients = []
    for email in recipients:
        domain = email.rsplit('@', 1)[-1].strip().lower() if '@' in email else ''
        if domain and domain in forced_domains:
            forced_smtp_recipients.append(email)
        else:
            regular_recipients.append(email)

    total_sent = 0
    total_failed = 0
    all_errors = []

    if forced_smtp_recipients:
        if smtp_configured():
            smtp_result = _send_email_via_smtp(forced_smtp_recipients, subject, plain_content, html_content)
            total_sent += smtp_result.get('sent', 0)
            total_failed += smtp_result.get('failed', 0)
            all_errors.extend(smtp_result.get('errors', []))
        else:
            # Keep delivery available in production even when SMTP isn't configured yet.
            # These recipients will be attempted via the provider/fallback path below.
            regular_recipients.extend(forced_smtp_recipients)

    if regular_recipients:
        if provider == 'brevo':
            try:
                result = _send_email_via_brevo_with_fallback(regular_recipients, subject, plain_content, html_content)
            except Exception as ex:
                if smtp_configured():
                    smtp_result = _send_email_via_smtp(regular_recipients, subject, plain_content, html_content)
                    total_sent += smtp_result.get('sent', 0)
                    total_failed += smtp_result.get('failed', 0)
                    all_errors.extend(smtp_result.get('errors', []))
                    if smtp_result.get('failed', 0) > 0:
                        all_errors.append({'email': '*', 'error': f'Brevo error before SMTP fallback: {str(ex)}'})
                else:
                    total_failed += len(regular_recipients)
                    all_errors.extend([
                        {'email': email, 'error': f'Brevo error and SMTP unavailable: {str(ex)}'}
                        for email in regular_recipients
                    ])
            else:
                total_sent += result.get('sent', 0)
                total_failed += result.get('failed', 0)
                all_errors.extend(result.get('errors', []))
        else:
            try:
                result = _send_email_via_smtp(regular_recipients, subject, plain_content, html_content)
            except Exception as ex:
                if brevo_configured():
                    brevo_result = _send_email_via_brevo(regular_recipients, subject, plain_content, html_content)
                    total_sent += brevo_result.get('sent', 0)
                    total_failed += brevo_result.get('failed', 0)
                    all_errors.extend(brevo_result.get('errors', []))
                    if brevo_result.get('failed', 0) > 0:
                        all_errors.append({'email': '*', 'error': f'SMTP error before Brevo fallback: {str(ex)}'})
                else:
                    total_failed += len(regular_recipients)
                    all_errors.extend([
                        {'email': email, 'error': f'SMTP error and Brevo unavailable: {str(ex)}'}
                        for email in regular_recipients
                    ])
            else:
                total_sent += result.get('sent', 0)
                total_failed += result.get('failed', 0)
                all_errors.extend(result.get('errors', []))

    return {'sent': total_sent, 'failed': total_failed, 'errors': all_errors}


def _send_email_via_brevo_with_fallback(recipients, subject, plain_content, html_content):
    if not brevo_configured():
        raise RuntimeError('Brevo is not configured. Set BREVO_API_KEY and SMTP_FROM_EMAIL in backend/.env.')

    result = _send_email_via_brevo(recipients, subject, plain_content, html_content)
    if result['failed'] == 0:
        return result

    # Keep delivery stable: if Brevo partially fails and SMTP is configured, retry only failed recipients via SMTP.
    if smtp_configured():
        failed_emails = [item['email'] for item in result['errors'] if item.get('email')]
        if failed_emails:
            smtp_result = _send_email_via_smtp(failed_emails, subject, plain_content, html_content)
            return {
                'sent': result['sent'] + smtp_result['sent'],
                'failed': smtp_result['failed'],
                'errors': smtp_result['errors'],
            }

    return result


def _send_email_via_brevo(recipients, subject, plain_content, html_content):
    sent_count = 0
    failed_count = 0
    errors = []
    timeout = max(5, int(config.BREVO_TIMEOUT or 20))
    plain_only_domains = {
        d.strip().lower()
        for d in (getattr(config, 'MAIL_PLAIN_ONLY_DOMAINS', None) or [])
        if d and d.strip()
    }
    custom_headers = _mail_metadata_headers()

    for email in recipients:
        domain = email.rsplit('@', 1)[-1].strip().lower() if '@' in email else ''
        plain_only = bool(domain and domain in plain_only_domains)

        payload = {
            'sender': {'name': config.SMTP_FROM_NAME, 'email': config.SMTP_FROM_EMAIL},
            'to': [{'email': email}],
            'subject': str(subject or ''),
            'textContent': plain_content,
            'replyTo': {'email': config.MAIL_REPLY_TO or config.SMTP_FROM_EMAIL},
            'headers': custom_headers,
            'tags': [config.MAIL_BULK_TAG] if config.MAIL_BULK_TAG else None,
        }
        if payload.get('tags') is None:
            payload.pop('tags', None)
        if not plain_only:
            payload['htmlContent'] = html_content
        data = json.dumps(payload).encode('utf-8')

        # Small retry window for transient Brevo API/network issues.
        sent = False
        last_error = 'unknown error'
        for attempt in range(3):
            req = urllib_request.Request(
                config.BREVO_API_URL,
                data=data,
                method='POST',
                headers={
                    'accept': 'application/json',
                    'content-type': 'application/json',
                    'api-key': config.BREVO_API_KEY,
                },
            )
            try:
                with urllib_request.urlopen(req, timeout=timeout) as resp:
                    status = getattr(resp, 'status', 0)
                    if 200 <= status < 300:
                        sent = True
                        break
                    last_error = f'Brevo returned status {status}'
            except urllib_error.HTTPError as ex:
                body = ex.read().decode('utf-8', errors='ignore') if hasattr(ex, 'read') else ''
                last_error = f'Brevo HTTP {ex.code}: {body[:220]}'
                if ex.code in {429, 500, 502, 503, 504} and attempt < 2:
                    time.sleep(1 + attempt)
                    continue
                break
            except urllib_error.URLError as ex:
                last_error = f'Brevo network error: {ex.reason}'
                if attempt < 2:
                    time.sleep(1 + attempt)
                    continue
                break
            except Exception as ex:
                last_error = str(ex)
                break

        if sent:
            sent_count += 1
        else:
            failed_count += 1
            errors.append({'email': email, 'error': last_error})

    return {'sent': sent_count, 'failed': failed_count, 'errors': errors}


def _send_email_via_smtp(recipients, subject, plain_content, html_content):
    if not smtp_configured():
        raise RuntimeError('SMTP is not configured. Set SMTP_* in backend/.env.')

    sent_count = 0
    failed_count = 0
    errors = []

    with smtplib.SMTP(config.SMTP_HOST, config.SMTP_PORT, timeout=20) as server:
        if config.SMTP_USE_TLS:
            server.starttls()
        server.login(config.SMTP_USERNAME, config.SMTP_PASSWORD)

        for email in recipients:
            msg = MIMEMultipart('alternative')
            msg['Subject'] = subject
            msg['From'] = f"{config.SMTP_FROM_NAME} <{config.SMTP_FROM_EMAIL}>"
            msg['To'] = email
            msg['Date'] = formatdate(localtime=True)
            msg['Message-ID'] = make_msgid(domain=(config.SMTP_FROM_EMAIL.split('@')[-1] if '@' in config.SMTP_FROM_EMAIL else None))
            if config.MAIL_REPLY_TO:
                msg['Reply-To'] = config.MAIL_REPLY_TO

            for key, value in _mail_metadata_headers().items():
                msg[key] = value

            msg.attach(MIMEText(plain_content, 'plain', 'utf-8'))
            msg.attach(MIMEText(html_content, 'html', 'utf-8'))
            try:
                server.sendmail(config.SMTP_FROM_EMAIL, [email], msg.as_string())
                sent_count += 1
            except Exception as ex:
                failed_count += 1
                errors.append({'email': email, 'error': str(ex)})

    return {'sent': sent_count, 'failed': failed_count, 'errors': errors}


def normalize_user_type(value):
    user_type = (value or '').strip().lower()
    if user_type in {'student', 'alumni'}:
        return user_type
    return None


def create_password_reset_otp(conn, person, user_type):
    otp = f"{random.randint(0, 999999):06d}"
    otp_hash = generate_password_hash(otp)
    expires_at = datetime.now() + timedelta(minutes=10)

    cur = conn.cursor()
    # Invalidate older active OTPs for the same mailbox and role.
    cur.execute(
        """
        UPDATE password_reset_otps
           SET used=1
         WHERE email=%s AND user_type=%s AND used=0
        """,
        (person['email'], user_type),
    )
    cur.execute(
        """
        INSERT INTO password_reset_otps (alumni_id, email, user_type, otp_hash, expires_at)
        VALUES (%s,%s,%s,%s,%s)
        """,
        (person['id'], person['email'], user_type, otp_hash, expires_at),
    )
    conn.commit()
    cur.close()
    return otp

# ─── Global JSON error handlers ───────────────────────
@app.errorhandler(RequestEntityTooLarge)
def handle_request_entity_too_large(_e):
    max_mb = max(1, int(app.config.get('MAX_CONTENT_LENGTH', 0) / (1024 * 1024)))
    return jsonify({
        'success': False,
        'message': f'Request payload too large. Please upload smaller images and keep total payload under {max_mb} MB.'
    }), 413


@app.errorhandler(Exception)
def handle_exception(e):
    """Return JSON for all unhandled exceptions instead of HTML."""
    if isinstance(e, HTTPException):
        message = (e.description or 'Request failed').strip()
        if e.code == 405:
            message = 'Method not allowed'
        if e.code == 413:
            max_mb = max(1, int(app.config.get('MAX_CONTENT_LENGTH', 0) / (1024 * 1024)))
            message = f'Request payload too large. Please upload smaller images and keep total payload under {max_mb} MB.'
        return jsonify({'success': False, 'message': message}), int(e.code or 500)

    if isinstance(e, pymysql_err.OperationalError):
        code = e.args[0] if e.args else None
        if code == 1045:
            return jsonify({
                'success': False,
                'message': 'Database authentication failed. Set MYSQL_USER and MYSQL_PASSWORD in backend/.env, then restart backend.'
            }), 503
        if code == 2003:
            return jsonify({
                'success': False,
                'message': 'Database unreachable. Verify MYSQL_HOST, MYSQL_PORT, and network allowlist in your DB provider.'
            }), 503
        if code == 2005:
            return jsonify({
                'success': False,
                'message': 'Database hostname is not resolvable. Check MYSQL_HOST or set MYSQL_URL from Aiven service URI.'
            }), 503
        if code == 2026:
            return jsonify({
                'success': False,
                'message': 'Database SSL error. For Aiven set MYSQL_SSL_MODE=required and configure MYSQL_SSL_CA if needed.'
            }), 503
        if code == 1049:
            return jsonify({
                'success': False,
                'message': 'Unknown database selected. Verify MYSQL_DB and run DB schema initialization.'
            }), 503

    if isinstance(e, pymysql_err.ProgrammingError):
        code = e.args[0] if e.args else None
        if code in {1146, 1054}:
            return jsonify({
                'success': False,
                'message': 'Database schema is not ready. Enable AUTO_INIT_DB or run schema.sql and migrate.py once.'
            }), 503

    if config.DEBUG:
        import traceback
        return jsonify({'success': False, 'message': str(e), 'trace': traceback.format_exc()[-500:]}), 500
    return jsonify({'success': False, 'message': 'Internal server error'}), 500

@app.errorhandler(404)
def not_found(e):
    return jsonify({'success': False, 'message': 'Not found'}), 404

# ─── DB connection helper ──────────────────────────────
def get_db():
    """Return a per-request cached DB connection (auto-closed at request teardown)."""
    db = g.get('db')
    if db is None:
        ensure_runtime_db_ready()
        g.db = _connect_mysql_with_retry(include_database=True, dict_cursor=True, autocommit=False)
        return g.db

    # Heal stale pooled connections between requests.
    try:
        db.ping(reconnect=True)
    except Exception:
        try:
            db.close()
        except Exception:
            pass
        g.db = _connect_mysql_with_retry(include_database=True, dict_cursor=True, autocommit=False)
    return g.db

@app.teardown_appcontext
def close_db(error):
    """Close the DB connection at the end of every request."""
    db = g.pop('db', None)
    if db is not None:
        try:
            db.close()
        except Exception:
            pass


@app.route('/uploads/<path:filename>', methods=['GET', 'OPTIONS'])
def serve_upload(filename):
    """Serve uploaded images with proper CORS/CORP headers and error handling"""
    # Handle CORS preflight
    if request.method == 'OPTIONS':
        resp = make_response('', 204)
        resp.headers['Access-Control-Allow-Origin'] = '*'
        resp.headers['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return resp

    try:
        # Sanitize filename to prevent directory traversal
        safe_filename = secure_filename(filename.split('/')[-1])
        if not safe_filename:
            return jsonify({'success': False, 'message': 'Invalid filename'}), 400

        # Try local filesystem first
        local_path = os.path.join(UPLOAD_FOLDER, filename)
        if os.path.isfile(local_path):
            resp = make_response(send_from_directory(UPLOAD_FOLDER, filename, mimetype=mimetypes.guess_type(filename)[0]))
            resp.headers['Access-Control-Allow-Origin'] = '*'
            resp.headers['Cross-Origin-Resource-Policy'] = 'cross-origin'
            return resp

        # Fallback to database
        row = fetch_uploaded_file(filename)
        if not row or not row.get('file_data'):
            return jsonify({'success': False, 'message': 'File not found'}), 404

        content_type = (row.get('content_type') or '').strip() or mimetypes.guess_type(filename)[0] or 'application/octet-stream'
        resp = make_response(send_file(
            BytesIO(row['file_data']),
            mimetype=content_type,
            as_attachment=False,
            download_name=filename,
        ))
        resp.headers['Access-Control-Allow-Origin'] = '*'
        resp.headers['Cross-Origin-Resource-Policy'] = 'cross-origin'
        return resp
    except Exception as e:
        print(f"[UPLOAD SERVE ERROR] {filename}: {e}")
        return jsonify({'success': False, 'message': f'Error serving file: {str(e)}'}), 500


@app.route('/api/uploads/<path:filename>', methods=['GET', 'OPTIONS'])
def serve_upload_api_alias(filename):
    """API alias for serving uploads - routes to main handler"""
    if request.method == 'OPTIONS':
        return '', 204
    # Reuse main handler but ensure CORS headers are present
    resp = serve_upload(filename)
    # If the response is a Flask response object, ensure CORS headers
    try:
        resp.headers['Access-Control-Allow-Origin'] = '*'
    except Exception:
        pass
    return resp


# ═══════════════════════════════════════════════════════
#  AUTH
# ═══════════════════════════════════════════════════════

@app.route('/api/admin-login', methods=['POST'])
def admin_login():
    try:
        data = request.get_json() or {}
        username = (data.get('username') or '').strip()
        password = (data.get('password') or '').strip()
        
        if not username or not password:
            return jsonify({'success': False, 'message': 'Username and password are required'}), 400
        
        if username == config.ADMIN_USERNAME and password == config.ADMIN_PASSWORD:
            return jsonify({'success': True, 'message': 'Admin login successful', 'admin': True}), 200
        
        return jsonify({'success': False, 'message': 'Invalid username or password'}), 401
    except Exception as e:
        print(f"[ADMIN LOGIN ERROR]: {e}")
        return jsonify({'success': False, 'message': 'Login error occurred'}), 500


@app.route('/api/alumni-login', methods=['POST'])
def alumni_login():
    data  = request.get_json()
    email = data.get('email', '').strip()
    pwd   = data.get('password', '')

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM alumni WHERE email = %s", (email,))
    alumni = cur.fetchone()
    cur.close()

    if not alumni:
        return jsonify({'success': False, 'message': 'Email not found'}), 404

    if alumni['status'] == 'pending':
        return jsonify({'success': False, 'message': 'Your registration is pending approval'}), 403

    if alumni['status'] == 'rejected':
        return jsonify({'success': False, 'message': 'Your registration was rejected'}), 403

    if not check_password_hash(alumni['password'], pwd):
        return jsonify({'success': False, 'message': 'Incorrect password'}), 401

    # Block students from logging in as alumni
    if (alumni.get('user_type') or 'alumni') == 'student':
        return jsonify({'success': False, 'message': 'This account is a student account. Please use the Student Login page.'}), 403

    past_jobs = fetch_past_jobs(conn, alumni['id'])

    return jsonify({
        'success': True,
        'alumni': {
            'id':          alumni['id'],
            'name':        alumni['name'],
            'email':       alumni['email'],
            'phone':       alumni['phone'],
            'department':  alumni['department'],
            'student_id':  alumni['student_id'],
            'session':     alumni['session'],
            'company':     alumni['company'],
            'designation': alumni['designation'],
            'current_job_start_date': str(alumni['current_job_start_date']) if alumni.get('current_job_start_date') else '',
            'past_jobs':          past_jobs,
            'status':             alumni['status'],
            'photo_url':          build_upload_url(alumni.get('photo')),
            'bio':                alumni.get('bio', ''),
            'research_interests': alumni.get('research_interests', ''),
            'extracurricular':    alumni.get('extracurricular', ''),
            'linkedin':           alumni.get('linkedin', ''),
            'github':             alumni.get('github', ''),
            'twitter':            alumni.get('twitter', ''),
            'website':            alumni.get('website', ''),
            'user_type':          alumni.get('user_type', 'alumni'),
            'upgrade_request':    alumni.get('upgrade_request'),
        }
    })


@app.route('/api/student-login', methods=['POST'])
def student_login():
    data  = request.get_json()
    email = data.get('email', '').strip()
    pwd   = data.get('password', '')

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM alumni WHERE email = %s", (email,))
    person = cur.fetchone()
    cur.close()

    if not person:
        return jsonify({'success': False, 'message': 'You are not registered yet. Please register as a student first.'}), 404

    # Upgraded alumni should not log in here
    if (person.get('user_type') or 'alumni') != 'student':
        if person.get('upgrade_request') == 'approved' or (person.get('user_type') or 'alumni') == 'alumni':
            return jsonify({'success': False, 'message': 'You have been upgraded to Alumni. Please use the Alumni Login page.'}), 403
        return jsonify({'success': False, 'message': 'You are not registered yet. Please register as a student first.'}), 403

    if person['status'] == 'pending':
        return jsonify({'success': False, 'message': 'Your registration is pending admin approval.'}), 403

    if person['status'] == 'rejected':
        return jsonify({'success': False, 'message': 'Your registration was rejected by the admin.'}), 403

    if not check_password_hash(person['password'], pwd):
        return jsonify({'success': False, 'message': 'Incorrect password.'}), 401

    return jsonify({
        'success': True,
        'alumni': {
            'id':             person['id'],
            'name':           person['name'],
            'email':          person['email'],
            'phone':          person['phone'],
            'department':     person['department'],
            'student_id':     person['student_id'],
            'session':        person['session'],
            'company':        person.get('company', ''),
            'designation':    person.get('designation', ''),
            'current_job_start_date': str(person['current_job_start_date']) if person.get('current_job_start_date') else '',
            'past_jobs':       fetch_past_jobs(conn, person['id']),
            'graduation_year':person.get('graduation_year', ''),
            'status':         person['status'],
            'user_type':      person.get('user_type', 'student'),
            'upgrade_request':person.get('upgrade_request'),
            'photo_url':      build_upload_url(person.get('photo')),
            'bio':            person.get('bio', ''),
        }
    })


@app.route('/api/forgot-password/request-otp', methods=['POST'])
def request_forgot_password_otp():
    data = request.get_json() or {}
    email = (data.get('email') or '').strip().lower()
    user_type = normalize_user_type(data.get('user_type'))

    if not email or not user_type:
        return jsonify({'success': False, 'message': 'Email and valid user type are required.'}), 400

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, name, email, user_type FROM alumni WHERE email=%s LIMIT 1", (email,))
    person = cur.fetchone()
    cur.close()

    # Do not expose account existence details.
    if not person:
        return jsonify({'success': True, 'message': 'If your account exists, an OTP has been sent to your email.'})

    person_type = normalize_user_type(person.get('user_type') or 'alumni')
    if person_type != user_type:
        return jsonify({'success': True, 'message': 'If your account exists, an OTP has been sent to your email.'})

    otp = create_password_reset_otp(conn, person, user_type)

    try:
        send_email(
            [person['email']],
            'AlumniConnect Password Reset OTP',
            (
                f"Hi {person.get('name') or 'there'},\n\n"
                f"Your password reset code is: {otp}\n"
                "This code is valid for 10 minutes.\n\n"
                "If you did not request this, you can ignore this email."
            ),
            'Use this OTP to reset your AlumniConnect password.',
            'Go to AlumniConnect',
        )
    except Exception as ex:
        return jsonify({'success': False, 'message': f'Failed to send OTP email: {ex}'}), 500

    return jsonify({'success': True, 'message': 'OTP sent to your email address.'})


@app.route('/api/forgot-password/reset', methods=['POST'])
def reset_password_with_otp():
    data = request.get_json() or {}
    email = (data.get('email') or '').strip().lower()
    user_type = normalize_user_type(data.get('user_type'))
    otp = (data.get('otp') or '').strip()
    new_password = data.get('new_password') or ''

    if not email or not user_type or not otp or not new_password:
        return jsonify({'success': False, 'message': 'Email, user type, OTP and new password are required.'}), 400

    if len(new_password) < 6:
        return jsonify({'success': False, 'message': 'New password must be at least 6 characters.'}), 400

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, name, email, user_type FROM alumni WHERE email=%s LIMIT 1", (email,))
    person = cur.fetchone()
    if not person:
        cur.close()
        return jsonify({'success': False, 'message': 'Account not found.'}), 404

    person_type = normalize_user_type(person.get('user_type') or 'alumni')
    if person_type != user_type:
        cur.close()
        return jsonify({'success': False, 'message': 'User type does not match this account.'}), 400

    cur.execute(
        """
        SELECT id, otp_hash, expires_at, attempts
        FROM password_reset_otps
        WHERE email=%s AND user_type=%s AND used=0
        ORDER BY id DESC
        LIMIT 1
        """,
        (email, user_type),
    )
    otp_row = cur.fetchone()

    if not otp_row:
        cur.close()
        return jsonify({'success': False, 'message': 'OTP not found. Please request a new one.'}), 400

    now = datetime.now()
    if otp_row['expires_at'] <= now:
        cur.execute("UPDATE password_reset_otps SET used=1 WHERE id=%s", (otp_row['id'],))
        conn.commit()
        cur.close()
        return jsonify({'success': False, 'message': 'OTP has expired. Please request a new one.'}), 400

    if int(otp_row.get('attempts') or 0) >= 5:
        cur.execute("UPDATE password_reset_otps SET used=1 WHERE id=%s", (otp_row['id'],))
        conn.commit()
        cur.close()
        return jsonify({'success': False, 'message': 'Too many failed attempts. Request a new OTP.'}), 429

    if not check_password_hash(otp_row['otp_hash'], otp):
        next_attempts = int(otp_row.get('attempts') or 0) + 1
        cur.execute(
            "UPDATE password_reset_otps SET attempts=%s, used=%s WHERE id=%s",
            (next_attempts, 1 if next_attempts >= 5 else 0, otp_row['id']),
        )
        conn.commit()
        cur.close()
        return jsonify({'success': False, 'message': 'Invalid OTP.'}), 400

    hashed_password = generate_password_hash(new_password)
    cur.execute("UPDATE alumni SET password=%s WHERE id=%s", (hashed_password, person['id']))
    cur.execute("UPDATE password_reset_otps SET used=1 WHERE id=%s", (otp_row['id'],))
    conn.commit()
    cur.close()

    try:
        send_email(
            [person['email']],
            'AlumniConnect Password Updated',
            (
                f"Hi {person.get('name') or 'there'},\n\n"
                "Your AlumniConnect password was successfully changed.\n"
                "If this was not you, please contact admin immediately."
            ),
            'Your password has been changed.',
            'Visit AlumniConnect',
        )
    except Exception:
        pass

    return jsonify({'success': True, 'message': 'Password reset successful. You can now log in.'})


# ═══════════════════════════════════════════════════════
#  ALUMNI REGISTRATION + MANAGEMENT
# ═══════════════════════════════════════════════════════

@app.route('/api/register', methods=['POST'])
def register():
    # Accept both multipart/form-data (with file) and JSON
    if request.content_type and 'multipart/form-data' in request.content_type:
        data = request.form
        photo_file = request.files.get('photo')
    else:
        data = request.get_json() or {}
        photo_file = None

    required = ['name', 'email', 'password']
    for f in required:
        if not data.get(f):
            return jsonify({'success': False, 'message': f'{f} is required'}), 400

    if len((data.get('password') or '')) < 6:
        return jsonify({'success': False, 'message': 'Password must be at least 6 characters.'}), 400

    # Save profile photo if provided
    photo_filename = save_uploaded_image(photo_file)
    if photo_file and photo_file.filename and not photo_filename:
        max_mb = max(1, int(MAX_IMAGE_UPLOAD_BYTES / (1024 * 1024)))
        return jsonify({'success': False, 'message': f'Photo is invalid or too large. Please upload JPG/PNG/WEBP/GIF up to {max_mb} MB.'}), 400

    # Save ID card photo if provided
    idcard_file = request.files.get('idcard') if (request.content_type and 'multipart/form-data' in request.content_type) else None
    idcard_filename = save_uploaded_image(idcard_file)
    if idcard_file and idcard_file.filename and not idcard_filename:
        max_mb = max(1, int(MAX_IMAGE_UPLOAD_BYTES / (1024 * 1024)))
        return jsonify({'success': False, 'message': f'ID card/evidence image is invalid or too large. Please upload JPG/PNG/WEBP/GIF up to {max_mb} MB.'}), 400

    email = (data.get('email') or '').strip()
    student_id = (data.get('student_id') or '').strip()
    hashed = generate_password_hash(data['password'])
    user_type = data.get('user_type', 'alumni')
    conn = get_db()
    cur = conn.cursor()
    try:
        # If an admin-added "Existing Alumni" row matches by email or student ID,
        # update that row as a real registration but keep it pending for admin approval.
        cur.execute(
            """
            SELECT id, is_manually_added
            FROM alumni
            WHERE is_manually_added=1
              AND (LOWER(email)=LOWER(%s) OR (%s <> '' AND student_id=%s))
            ORDER BY id DESC
            LIMIT 1
            """,
            (email, student_id, student_id),
        )
        existing = cur.fetchone()
        if existing:
            cur.execute(
                """
                UPDATE alumni
                   SET name=%s,
                       email=%s,
                       phone=%s,
                       department=%s,
                       student_id=%s,
                       session=%s,
                       graduation_year=%s,
                       hall_name=%s,
                       company=%s,
                       designation=%s,
                       password=%s,
                       photo=CASE WHEN %s IS NULL THEN photo ELSE %s END,
                       id_photo=CASE WHEN %s IS NULL THEN id_photo ELSE %s END,
                                             status='pending',
                       user_type=%s,
                                             is_manually_added=1,
                       upgrade_request=NULL,
                       upgrade_document=NULL
                 WHERE id=%s
                """,
                (
                    data.get('name'), email, data.get('phone'),
                    data.get('department', 'ICE'), data.get('student_id'),
                    data.get('session'), data.get('graduation_year'), data.get('hall_name'),
                    data.get('company'), data.get('designation'),
                    hashed,
                    photo_filename, photo_filename,
                    idcard_filename, idcard_filename,
                    user_type,
                    existing['id'],
                )
            )
            conn.commit()
            warning = None
            try:
                send_registration_received_email(data.get('name'), email, user_type)
            except Exception as ex:
                warning = str(ex)
            return jsonify({
                'success': True,
                'id': existing['id'],
                'message': 'Registration submitted. Waiting for admin approval.',
                'email_warning': warning,
            }), 201

        cur.execute(
            """INSERT INTO alumni (name, email, phone, department, student_id, session,
                                  graduation_year, hall_name, company, designation, password, photo, id_photo, status, user_type)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'pending',%s)""",
            (data.get('name'), email, data.get('phone'),
             data.get('department', 'ICE'), data.get('student_id'),
             data.get('session'), data.get('graduation_year'), data.get('hall_name'),
             data.get('company'), data.get('designation'),
             hashed, photo_filename, idcard_filename, user_type)
        )
        conn.commit()
        new_id = cur.lastrowid
    except Exception as e:
        conn.rollback()
        code = e.args[0] if getattr(e, 'args', None) else None
        if 'Duplicate entry' in str(e) or code == 1062:
            return jsonify({'success': False, 'message': 'Email already registered'}), 409
        if code in {1146, 1054}:
            return jsonify({'success': False, 'message': 'Database tables are not initialized yet. Please contact admin.'}), 503
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        cur.close()

    warning = None
    try:
        send_registration_received_email(data.get('name'), email, user_type)
    except Exception as ex:
        warning = str(ex)

    return jsonify({
        'success': True,
        'id': new_id,
        'message': 'Registration submitted. Waiting for admin approval.',
        'email_warning': warning,
    }), 201


@app.route('/api/pending', methods=['GET'])
def get_pending():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id,name,email,phone,department,student_id,session,graduation_year,company,designation,photo,id_photo,status,user_type,created_at FROM alumni WHERE status='pending' ORDER BY created_at DESC")
    rows = cur.fetchall()
    cur.close()
    for r in rows:
        r['photo_url'] = build_upload_url(r.get('photo'))
        r['id_photo_url'] = build_upload_url(r.get('id_photo'))
    return jsonify(rows)


@app.route('/api/alumni', methods=['GET'])
def get_alumni():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id,name,email,phone,address,department,student_id,session,graduation_year,company,designation,current_job_start_date,higher_study,photo,bio,research_interests,extracurricular,linkedin,github,twitter,website,status,created_at FROM alumni WHERE status='approved' AND (user_type IS NULL OR user_type='alumni') AND (is_manually_added IS NULL OR is_manually_added=0) ORDER BY name")
    rows = cur.fetchall()

    cur.execute(
        """
        SELECT alumni_id, id, company, designation, start_date, end_date, created_at
        FROM past_job_experiences
        ORDER BY COALESCE(end_date, '9999-12-31') DESC, id DESC
        """
    )
    past_rows = cur.fetchall()
    cur.close()

    past_jobs_by_alumni = {}
    for p in past_rows:
        p_item = {
            'id': p.get('id'),
            'company': p.get('company'),
            'designation': p.get('designation'),
            'start_date': str(p['start_date']) if p.get('start_date') else None,
            'end_date': str(p['end_date']) if p.get('end_date') else None,
            'created_at': str(p['created_at']) if p.get('created_at') else None,
        }
        past_jobs_by_alumni.setdefault(p['alumni_id'], []).append(p_item)

    for r in rows:
        r['photo_url'] = build_upload_url(r.get('photo'))
        if r.get('current_job_start_date'):
            r['current_job_start_date'] = str(r['current_job_start_date'])
        r['past_jobs'] = past_jobs_by_alumni.get(r['id'], [])
    return jsonify(rows)


@app.route('/api/students', methods=['GET'])
def get_students():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id,name,email,phone,address,department,student_id,session,company,designation,current_job_start_date,higher_study,bio,research_interests,extracurricular,linkedin,github,twitter,website,photo,status,created_at FROM alumni WHERE status='approved' AND user_type='student' ORDER BY name")
    rows = cur.fetchall()
    cur.close()
    for r in rows:
        r['photo_url'] = build_upload_url(r.get('photo'))
    return jsonify(rows)


PROFILE_SELECT_SQL = "SELECT id,name,email,phone,address,department,student_id,session,graduation_year,hall_name,company,designation,current_job_start_date,higher_study,photo,bio,research_interests,extracurricular,linkedin,github,twitter,website,status,user_type,upgrade_request,created_at FROM alumni WHERE id=%s"
PROFILE_EDIT_FIELDS = ['name','email','phone','address','department','student_id','session','graduation_year','hall_name','company','designation','current_job_start_date','higher_study','bio','research_interests','extracurricular','linkedin','github','twitter','website']


def _serialize_profile(conn, row):
    if not row:
        return None

    profile = dict(row)
    profile['photo_url'] = build_upload_url(profile.get('photo'))
    profile['verified'] = profile.get('status') == 'approved' and (profile.get('user_type') or 'alumni') != 'student'
    profile['role_label'] = 'Student' if profile.get('user_type') == 'student' else 'Alumni'

    if profile.get('current_job_start_date'):
        profile['current_job_start_date'] = str(profile['current_job_start_date'])
    if profile.get('created_at'):
        profile['created_at'] = str(profile['created_at'])

    profile['past_jobs'] = fetch_past_jobs(conn, profile['id'])
    return profile


def _get_profile_row(conn, pid):
    cur = conn.cursor()
    try:
        cur.execute(PROFILE_SELECT_SQL, (pid,))
        return cur.fetchone()
    finally:
        cur.close()


def _save_profile_updates(conn, pid, data, photo_file=None):
    fields = [field for field in PROFILE_EDIT_FIELDS if field in data]
    has_past_jobs = 'past_jobs' in data

    if photo_file is None and not fields and not has_past_jobs:
        return jsonify({'success': False, 'message': 'No fields to update'}), 400

    cur = conn.cursor()
    try:
        cur.execute("SELECT company, designation, current_job_start_date FROM alumni WHERE id=%s", (pid,))
        existing = cur.fetchone()
        if not existing:
            return jsonify({'success': False, 'message': 'Profile not found'}), 404

        prev_company = (existing.get('company') or '').strip()
        prev_designation = (existing.get('designation') or '').strip()
        prev_start = existing.get('current_job_start_date')

        next_company = (data['company'] if 'company' in data else existing.get('company') or '').strip()
        next_designation = (data['designation'] if 'designation' in data else existing.get('designation') or '').strip()
        next_start = data['current_job_start_date'] if 'current_job_start_date' in data else existing.get('current_job_start_date')

        current_job_changed = (next_company != prev_company) or (next_designation != prev_designation)
        update_values = {}
        for field in fields:
            value = data.get(field)
            update_values[field] = None if value in ('', None) else value

        if photo_file is not None:
            filename = save_uploaded_image(photo_file)
            if not filename:
                return jsonify({'success': False, 'message': 'Invalid image file'}), 400
            update_values['photo'] = filename

        if update_values:
            sets = ', '.join(f"{field}=%s" for field in update_values.keys())
            cur.execute(
                f"UPDATE alumni SET {sets} WHERE id=%s",
                [*update_values.values(), pid],
            )

        if has_past_jobs:
            provided = data.get('past_jobs') or []
            if not isinstance(provided, list):
                return jsonify({'success': False, 'message': 'past_jobs must be a list'}), 400

            cur.execute("DELETE FROM past_job_experiences WHERE alumni_id=%s", (pid,))
            for item in provided:
                if not isinstance(item, dict):
                    continue
                company = (item.get('company') or '').strip()
                designation = (item.get('designation') or '').strip()
                start_date = item.get('start_date') or None
                end_date = item.get('end_date') or None
                if not company and not designation:
                    continue
                cur.execute(
                    "INSERT INTO past_job_experiences (alumni_id, company, designation, start_date, end_date) VALUES (%s,%s,%s,%s,%s)",
                    (pid, company or None, designation or None, start_date, end_date),
                )
        elif current_job_changed and (prev_company or prev_designation):
            auto_end = next_start or date.today().isoformat()
            cur.execute(
                """
                SELECT id
                FROM past_job_experiences
                WHERE alumni_id=%s
                  AND COALESCE(company,'')=%s
                  AND COALESCE(designation,'')=%s
                  AND COALESCE(start_date,'1000-01-01') = COALESCE(%s,'1000-01-01')
                  AND COALESCE(end_date,'1000-01-01') = COALESCE(%s,'1000-01-01')
                LIMIT 1
                """,
                (pid, prev_company, prev_designation, prev_start, auto_end),
            )
            if not cur.fetchone():
                cur.execute(
                    "INSERT INTO past_job_experiences (alumni_id, company, designation, start_date, end_date) VALUES (%s,%s,%s,%s,%s)",
                    (pid, prev_company or None, prev_designation or None, prev_start, auto_end),
                )

        conn.commit()
        updated = _get_profile_row(conn, pid)
        return jsonify({'success': True, 'profile': _serialize_profile(conn, updated)})
    except Exception as ex:
        conn.rollback()
        return jsonify({'success': False, 'message': str(ex)}), 500
    finally:
        cur.close()


@app.route('/api/profile/<int:pid>', methods=['GET'])
def get_profile(pid):
    conn = get_db()
    row = _get_profile_row(conn, pid)
    if not row:
        return jsonify({'success': False, 'message': 'Profile not found'}), 404
    return jsonify({'success': True, 'profile': _serialize_profile(conn, row)})


@app.route('/api/my-profile', methods=['GET'])
def get_my_profile():
    pid = request.args.get('id') or request.args.get('user_id')
    user_type = (request.args.get('user_type') or '').strip()
    if not pid:
        return jsonify({'success': False, 'message': 'user id is required'}), 400

    conn = get_db()
    row = _get_profile_row(conn, pid)
    if not row:
        return jsonify({'success': False, 'message': 'Profile not found'}), 404
    if user_type and (row.get('user_type') or 'alumni') != user_type:
        return jsonify({'success': False, 'message': 'Profile type mismatch'}), 403
    return jsonify({'success': True, 'profile': _serialize_profile(conn, row)})


@app.route('/api/edit-profile', methods=['PUT'])
def edit_profile():
    if request.form:
        data = request.form.to_dict(flat=True)
    else:
        data = request.get_json() or {}

    pid = data.get('user_id') or data.get('id')
    if not pid:
        return jsonify({'success': False, 'message': 'user_id is required'}), 400

    conn = get_db()
    photo_file = request.files.get('photo')
    return _save_profile_updates(conn, pid, data, photo_file=photo_file)


@app.route('/api/alumni/<int:aid>', methods=['PUT'])
def update_alumni(aid):
    data = request.get_json() or {}
    fields = ['name','phone','address','company','designation','current_job_start_date','higher_study','session','hall_name','bio','research_interests','extracurricular','linkedin','github','twitter','website']
    update_fields = [f for f in fields if f in data]
    has_past_jobs = 'past_jobs' in data

    if not update_fields and not has_past_jobs:
        return jsonify({'success': False, 'message': 'No fields to update'}), 400

    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("SELECT company, designation, current_job_start_date FROM alumni WHERE id=%s", (aid,))
        existing = cur.fetchone()
        if not existing:
            return jsonify({'success': False, 'message': 'Alumni not found'}), 404

        prev_company = (existing.get('company') or '').strip()
        prev_designation = (existing.get('designation') or '').strip()
        prev_start = existing.get('current_job_start_date')

        next_company = (data['company'] if 'company' in data else existing.get('company') or '').strip()
        next_designation = (data['designation'] if 'designation' in data else existing.get('designation') or '').strip()
        next_start = data['current_job_start_date'] if 'current_job_start_date' in data else existing.get('current_job_start_date')

        current_job_changed = (next_company != prev_company) or (next_designation != prev_designation)

        if update_fields:
            sets = ', '.join(f"{f}=%s" for f in update_fields)
            vals = [data[f] for f in update_fields]
            vals.append(aid)
            cur.execute(f"UPDATE alumni SET {sets} WHERE id=%s", vals)

        if has_past_jobs:
            provided = data.get('past_jobs') or []
            if not isinstance(provided, list):
                return jsonify({'success': False, 'message': 'past_jobs must be a list'}), 400

            cur.execute("DELETE FROM past_job_experiences WHERE alumni_id=%s", (aid,))
            for item in provided:
                if not isinstance(item, dict):
                    continue
                company = (item.get('company') or '').strip()
                designation = (item.get('designation') or '').strip()
                start_date = item.get('start_date') or None
                end_date = item.get('end_date') or None
                if not company and not designation:
                    continue
                cur.execute(
                    "INSERT INTO past_job_experiences (alumni_id, company, designation, start_date, end_date) VALUES (%s,%s,%s,%s,%s)",
                    (aid, company or None, designation or None, start_date, end_date),
                )
        elif current_job_changed and (prev_company or prev_designation):
            auto_end = next_start or date.today().isoformat()
            cur.execute(
                """
                SELECT id
                FROM past_job_experiences
                WHERE alumni_id=%s
                  AND COALESCE(company,'')=%s
                  AND COALESCE(designation,'')=%s
                  AND COALESCE(start_date,'1000-01-01') = COALESCE(%s,'1000-01-01')
                  AND COALESCE(end_date,'1000-01-01') = COALESCE(%s,'1000-01-01')
                LIMIT 1
                """,
                (aid, prev_company, prev_designation, prev_start, auto_end),
            )
            if not cur.fetchone():
                cur.execute(
                    "INSERT INTO past_job_experiences (alumni_id, company, designation, start_date, end_date) VALUES (%s,%s,%s,%s,%s)",
                    (aid, prev_company or None, prev_designation or None, prev_start, auto_end),
                )

        conn.commit()
        updated = _get_profile_row(conn, aid)
        return jsonify({'success': True, 'profile': _serialize_profile(conn, updated)})
    except Exception as ex:
        conn.rollback()
        return jsonify({'success': False, 'message': str(ex)}), 500
    finally:
        cur.close()


@app.route('/api/alumni/<int:aid>', methods=['DELETE'])
def delete_alumni(aid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM alumni WHERE id=%s", (aid,))
    conn.commit()
    cur.close()
    return jsonify({'success': True})


@app.route('/api/approve/<int:aid>', methods=['POST'])
def approve(aid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT name, email, user_type FROM alumni WHERE id=%s", (aid,))
    person = cur.fetchone()
    # On approval, clear manual flag so user leaves Existing Alumni list and appears in All Alumni.
    cur.execute("UPDATE alumni SET status='approved', is_manually_added=0 WHERE id=%s", (aid,))
    conn.commit()
    cur.close()

    warning = None
    if person and person.get('email'):
        try:
            role_label = 'student' if (person.get('user_type') == 'student') else 'alumni'
            send_email(
                [person['email']],
                'Your AlumniConnect registration has been approved',
                f"Hi {person.get('name') or 'there'},\n\nYour {role_label} registration has been approved by admin. You can now login and use your dashboard.",
                'Congratulations. Your account is now active.',
                'Login to AlumniConnect'
            )
        except Exception as ex:
            warning = str(ex)

    return jsonify({'success': True, 'email_warning': warning})


@app.route('/api/reject/<int:aid>', methods=['POST'])
def reject(aid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE alumni SET status='rejected' WHERE id=%s", (aid,))
    conn.commit()
    cur.close()
    return jsonify({'success': True})


# ═══════════════════════════════════════════════════════
#  ALUMNI REFERRALS
# ═══════════════════════════════════════════════════════

@app.route('/api/referrals', methods=['POST'])
def create_referral():
    data = request.get_json() or {}
    referred_by_alumni_id = data.get('referred_by_alumni_id')
    referred_name = (data.get('referred_name') or '').strip()
    referred_email = (data.get('referred_email') or '').strip().lower()
    referred_phone = (data.get('referred_phone') or '').strip() or None
    referred_student_id = (data.get('referred_student_id') or '').strip() or None
    referred_session = (data.get('referred_session') or '').strip() or None
    referred_department = (data.get('referred_department') or 'ICE').strip() or 'ICE'
    relation_note = (data.get('relation_note') or '').strip() or None

    if not referred_by_alumni_id or not referred_name or not referred_email:
        return jsonify({'success': False, 'message': 'referred_by_alumni_id, referred_name and referred_email are required.'}), 400

    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT id FROM alumni
            WHERE id=%s AND status='approved' AND (user_type IS NULL OR user_type='alumni')
            LIMIT 1
            """,
            (referred_by_alumni_id,),
        )
        referrer = cur.fetchone()
        if not referrer:
            return jsonify({'success': False, 'message': 'Only approved alumni can submit referrals.'}), 403

        cur.execute(
            """
            SELECT id
            FROM alumni_referrals
            WHERE status='pending'
              AND (LOWER(referred_email)=LOWER(%s) OR (%s IS NOT NULL AND referred_student_id=%s))
            LIMIT 1
            """,
            (referred_email, referred_student_id, referred_student_id),
        )
        if cur.fetchone():
            return jsonify({'success': False, 'message': 'A pending referral already exists for this person.'}), 409

        cur.execute(
            """
            INSERT INTO alumni_referrals
            (referred_by_alumni_id, referred_name, referred_email, referred_phone, referred_student_id,
             referred_session, referred_department, relation_note, status)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'pending')
            """,
            (
                referred_by_alumni_id,
                referred_name,
                referred_email,
                referred_phone,
                referred_student_id,
                referred_session,
                referred_department,
                relation_note,
            ),
        )
        conn.commit()
        referral_id = cur.lastrowid
        return jsonify({'success': True, 'id': referral_id, 'message': 'Referral submitted for admin approval.'}), 201
    except Exception as ex:
        conn.rollback()
        return jsonify({'success': False, 'message': str(ex)}), 500
    finally:
        cur.close()


@app.route('/api/referrals', methods=['GET'])
def get_referrals_for_alumni():
    alumni_id = request.args.get('alumni_id', type=int)
    if not alumni_id:
        return jsonify({'success': False, 'message': 'alumni_id is required.'}), 400

    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, referred_name, referred_email, referred_phone, referred_student_id,
               referred_session, referred_department, relation_note, status, admin_note,
               reviewed_at, created_at
        FROM alumni_referrals
        WHERE referred_by_alumni_id=%s
        ORDER BY created_at DESC, id DESC
        """,
        (alumni_id,),
    )
    rows = cur.fetchall()
    cur.close()
    for r in rows:
        if r.get('created_at'):
            r['created_at'] = str(r['created_at'])
        if r.get('reviewed_at'):
            r['reviewed_at'] = str(r['reviewed_at'])
    return jsonify(rows)


@app.route('/api/referrals/pending', methods=['GET'])
def get_pending_referrals():
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT r.id, r.referred_name, r.referred_email, r.referred_phone,
               r.referred_student_id, r.referred_session, r.referred_department,
               r.relation_note, r.status, r.created_at,
               a.id AS referrer_id, a.name AS referrer_name, a.email AS referrer_email
        FROM alumni_referrals r
        LEFT JOIN alumni a ON a.id = r.referred_by_alumni_id
        WHERE r.status='pending'
        ORDER BY r.created_at DESC, r.id DESC
        """
    )
    rows = cur.fetchall()
    cur.close()
    for r in rows:
        if r.get('created_at'):
            r['created_at'] = str(r['created_at'])
    return jsonify(rows)


@app.route('/api/referrals/<int:referral_id>/approve', methods=['POST'])
def approve_referral(referral_id):
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT id, referred_name, referred_email, referred_phone, referred_student_id,
                   referred_session, referred_department
            FROM alumni_referrals
            WHERE id=%s AND status='pending'
            LIMIT 1
            """,
            (referral_id,),
        )
        referral = cur.fetchone()
        if not referral:
            return jsonify({'success': False, 'message': 'Pending referral not found.'}), 404

        student_id = (referral.get('referred_student_id') or '').strip()
        cur.execute(
            """
            SELECT id, is_manually_added
            FROM alumni
            WHERE LOWER(email)=LOWER(%s) OR (%s <> '' AND student_id=%s)
            ORDER BY id DESC
            LIMIT 1
            """,
            (referral['referred_email'], student_id, student_id),
        )
        existing = cur.fetchone()

        if existing:
            cur.execute(
                """
                UPDATE alumni
                SET name=%s,
                    email=%s,
                    phone=%s,
                    student_id=%s,
                    session=%s,
                    department=%s,
                    status='approved',
                    user_type='alumni',
                    is_manually_added=1
                WHERE id=%s
                """,
                (
                    referral['referred_name'],
                    referral['referred_email'],
                    referral.get('referred_phone'),
                    referral.get('referred_student_id'),
                    referral.get('referred_session'),
                    referral.get('referred_department') or 'ICE',
                    existing['id'],
                ),
            )
        else:
            dummy_password = generate_password_hash('ADMIN_ADDED_' + str(uuid.uuid4())[:8])
            cur.execute(
                """
                INSERT INTO alumni (name, email, phone, student_id, session, department,
                                    password, status, user_type, is_manually_added)
                VALUES (%s,%s,%s,%s,%s,%s,%s,'approved','alumni',1)
                """,
                (
                    referral['referred_name'],
                    referral['referred_email'],
                    referral.get('referred_phone'),
                    referral.get('referred_student_id'),
                    referral.get('referred_session'),
                    referral.get('referred_department') or 'ICE',
                    dummy_password,
                ),
            )

        cur.execute(
            """
            UPDATE alumni_referrals
            SET status='approved', reviewed_by='admin', reviewed_at=NOW()
            WHERE id=%s
            """,
            (referral_id,),
        )
        conn.commit()
    except Exception as ex:
        conn.rollback()
        return jsonify({'success': False, 'message': str(ex)}), 500
    finally:
        cur.close()

    warning = None
    try:
        send_referral_invitation_email(referral.get('referred_name'), referral.get('referred_email'))
    except Exception as ex:
        warning = str(ex)

    return jsonify({'success': True, 'email_warning': warning})


@app.route('/api/referrals/<int:referral_id>/reject', methods=['POST'])
def reject_referral(referral_id):
    data = request.get_json(silent=True) or {}
    admin_note = (data.get('admin_note') or '').strip() or None

    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE alumni_referrals
        SET status='rejected', admin_note=%s, reviewed_by='admin', reviewed_at=NOW()
        WHERE id=%s AND status='pending'
        """,
        (admin_note, referral_id),
    )
    changed = cur.rowcount
    conn.commit()
    cur.close()

    if changed == 0:
        return jsonify({'success': False, 'message': 'Pending referral not found.'}), 404
    return jsonify({'success': True})


@app.route('/api/request-upgrade/<int:aid>', methods=['POST'])
def request_upgrade(aid):
    content_type = (request.content_type or '').lower()
    is_multipart = ('multipart/form-data' in content_type) or bool(request.files)
    data = request.form if is_multipart else (request.get_json(silent=True) or {})
    document_file = request.files.get('document') if is_multipart else None

    conn = get_db()
    cur = conn.cursor()

    document_filename = None
    if document_file and document_file.filename:
        ext = document_file.filename.rsplit('.', 1)[-1].lower()
        if ext not in ALLOWED_DOCUMENT_EXTENSIONS:
            cur.close()
            return jsonify({'success': False, 'message': 'Invalid file type. Upload JPG, JPEG, PNG, or WEBP image only.'}), 400
        document_filename = save_uploaded_image(document_file)
        if not document_filename:
            cur.close()
            max_mb = max(1, int(MAX_IMAGE_UPLOAD_BYTES / (1024 * 1024)))
            return jsonify({'success': False, 'message': f'Uploaded document is too large. Keep file size up to {max_mb} MB.'}), 400

    # Optionally update graduation details provided at request time
    updates = []
    vals = []
    for field in ['graduation_year', 'company', 'designation']:
        if data.get(field):
            updates.append(f"{field}=%s")
            vals.append(data[field])
    if document_filename:
        updates.append("upgrade_document=%s")
        vals.append(document_filename)
    updates.append("upgrade_request='pending'")
    vals.append(aid)
    cur.execute(f"UPDATE alumni SET {', '.join(updates)} WHERE id=%s", vals)
    conn.commit()
    cur.close()
    return jsonify({'success': True})


@app.route('/api/upgrade-requests', methods=['GET'])
def get_upgrade_requests():
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """SELECT id,name,email,phone,department,student_id,session,graduation_year,
                company,designation,photo,id_photo,upgrade_document,status,user_type,upgrade_request,created_at
           FROM alumni WHERE upgrade_request='pending' ORDER BY created_at DESC"""
    )
    rows = cur.fetchall()
    cur.close()
    for r in rows:
        r['photo_url'] = build_upload_url(r.get('photo'))
        r['id_photo_url'] = build_upload_url(r.get('id_photo'))
        r['upgrade_document_url'] = build_upload_url(r.get('upgrade_document'))
    return jsonify(rows)


@app.route('/api/approve-upgrade/<int:aid>', methods=['POST'])
def approve_upgrade(aid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT name, email FROM alumni WHERE id=%s", (aid,))
    person = cur.fetchone()
    cur.execute("UPDATE alumni SET user_type='alumni', upgrade_request='approved' WHERE id=%s", (aid,))
    conn.commit()
    cur.close()

    warning = None
    if person and person.get('email'):
        try:
            send_email(
                [person['email']],
                'Your Alumni upgrade request has been approved',
                f"Hi {person.get('name') or 'there'},\n\nGreat news. Your account has been upgraded to Alumni status. Please login from the Alumni Login page.",
                'Your membership tier has been upgraded.',
                'Open Alumni Dashboard'
            )
        except Exception as ex:
            warning = str(ex)

    return jsonify({'success': True, 'email_warning': warning})


@app.route('/api/reject-upgrade/<int:aid>', methods=['POST'])
def reject_upgrade(aid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE alumni SET upgrade_request='rejected' WHERE id=%s", (aid,))
    conn.commit()
    cur.close()
    return jsonify({'success': True})


# ═══════════════════════════════════════════════════════
#  EVENTS
# ═══════════════════════════════════════════════════════

@app.route('/api/events', methods=['GET'])
def get_events():
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("SELECT * FROM events ORDER BY created_at DESC, id DESC")
    except pymysql_err.OperationalError as e:
        # Fallback for old schema that does not yet have created_at.
        if e.args and e.args[0] == 1054:
            cur.execute("SELECT * FROM events ORDER BY id DESC")
        else:
            raise
    rows = cur.fetchall()
    cur.close()
    # stringify dates and build image URLs
    for r in rows:
        if r.get('date'):
            r['date'] = str(r['date'])
        # Ensure registration_deadline is serialized as string for clients
        if r.get('registration_deadline'):
            r['registration_deadline'] = str(r['registration_deadline'])
        # MySQL TIME can come as timedelta; normalize for JSON response.
        if r.get('event_time') is not None:
            r['event_time'] = str(r['event_time'])
            if not r.get('time'):
                r['time'] = r['event_time']
        if r.get('created_at'):
            r['created_at'] = str(r['created_at'])
        # Build banner image URL from stored filename
        if r.get('banner_image'):
            r['banner_image_url'] = build_upload_url(r.get('banner_image'))
    return jsonify(rows)


@app.route('/api/events', methods=['POST'])
def add_event():
    content_type = (request.content_type or '').lower()
    is_multipart = ('multipart/form-data' in content_type) or bool(request.files)
    data = request.form if is_multipart else (request.get_json(silent=True) or {})

    title = (data.get('title') or '').strip()
    date_value = (data.get('date') or '').strip()
    location = (data.get('location') or '').strip()
    description = (data.get('description') or '').strip()
    time_value = (data.get('time') or data.get('event_time') or '').strip() or None
    registration_deadline = (data.get('registration_deadline') or '').strip() or None
    payment_account = (data.get('payment_account') or '').strip() or None
    audience = (data.get('audience') or 'both').strip().lower()
    if audience not in ('both', 'alumni', 'students'):
        audience = 'both'

    if not title:
        return jsonify({'success': False, 'message': 'title is required'}), 400
    if not date_value:
        return jsonify({'success': False, 'message': 'date is required'}), 400
    if not location:
        return jsonify({'success': False, 'message': 'location is required'}), 400

    try:
        datetime.strptime(date_value, '%Y-%m-%d')
    except ValueError:
        return jsonify({'success': False, 'message': 'date must be YYYY-MM-DD'}), 400

    if registration_deadline:
        try:
            datetime.strptime(registration_deadline, '%Y-%m-%d')
        except ValueError:
            return jsonify({'success': False, 'message': 'registration_deadline must be YYYY-MM-DD'}), 400

    if time_value:
        try:
            datetime.strptime(time_value, '%H:%M')
        except ValueError:
            try:
                datetime.strptime(time_value, '%H:%M:%S')
            except ValueError:
                return jsonify({'success': False, 'message': 'time must be HH:MM or HH:MM:SS'}), 400

    fee_raw = data.get('fee', 0)
    try:
        fee_value = float(fee_raw or 0)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': 'fee must be a valid number'}), 400
    if fee_value < 0:
        return jsonify({'success': False, 'message': 'fee cannot be negative'}), 400

    banner_file = request.files.get('banner_image') or request.files.get('file')
    banner_image = None
    if banner_file and banner_file.filename:
        stored_banner = save_uploaded_image(banner_file)
        if not stored_banner:
            max_mb = max(1, int(MAX_IMAGE_UPLOAD_BYTES / (1024 * 1024)))
            allowed_types = ', '.join(sorted(ALLOWED_EXTENSIONS))
            return jsonify({'success': False, 'message': f'Invalid banner image file. Allowed types: {allowed_types}. Max size: {max_mb}MB.'}), 400
        banner_image = build_upload_url(stored_banner)

    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO events (title, date, event_time, registration_deadline, location, description, fee, payment_account, audience, banner_image) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (title, date_value, time_value, registration_deadline, location, description,
             fee_value, payment_account, audience, banner_image)
        )
    except pymysql_err.OperationalError as e:
        if e.args and e.args[0] == 1054:  # Unknown column: keep compatibility with older schema
            try:
                cur.execute(
                    "INSERT INTO events (title, date, event_time, location, description, fee, payment_account, audience, banner_image) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    (title, date_value, time_value, location, description, fee_value, payment_account, audience, banner_image)
                )
            except pymysql_err.OperationalError as e2:
                if e2.args and e2.args[0] == 1054:
                    cur.execute(
                        "INSERT INTO events (title, date, location, description, fee, payment_account, audience) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                        (title, date_value, location, description, fee_value, payment_account, audience)
                    )
                else:
                    raise
        else:
            raise
    conn.commit()
    new_id = cur.lastrowid

    if audience == 'students':
        cur.execute("SELECT email FROM alumni WHERE status='approved' AND user_type='student' AND email IS NOT NULL AND email <> ''")
    elif audience == 'alumni':
        cur.execute("SELECT email FROM alumni WHERE status='approved' AND (user_type IS NULL OR user_type='alumni') AND email IS NOT NULL AND email <> ''")
    else:
        cur.execute("SELECT email FROM alumni WHERE status='approved' AND email IS NOT NULL AND email <> ''")
    event_recipients = [r['email'] for r in cur.fetchall()]

    cur.close()

    warning = None
    try:
        send_email(
            event_recipients,
            f"New Event Notice: {title or 'AlumniConnect Event'}",
            f"A new event has been published.\n\nTitle: {title}\nDate: {date_value}\nTime: {time_value or ''}\nLocation: {location}\n\n{description}",
            'A new event has been announced for your community.',
            'View Event Details'
        )
    except Exception as ex:
        warning = str(ex)

    created_event = {
        'id': new_id,
        'title': title,
        'date': date_value,
        'time': time_value,
        'event_time': time_value,
        'registration_deadline': registration_deadline,
        'location': location,
        'description': description,
        'fee': fee_value,
        'payment_account': payment_account,
        'audience': audience,
        'banner_image': banner_image,
        'banner_image_url': banner_image,
    }

    return jsonify({'success': True, 'id': new_id, 'event': created_event, 'email_warning': warning}), 201


@app.route('/api/email/send', methods=['POST'])
def send_bulk_email_endpoint():
    data = request.get_json() or {}
    recipients = _normalize_recipient_emails(data.get('recipient_emails') or [])
    subject = (data.get('subject') or '').strip()
    message = (data.get('message') or '').strip()
    preheader = (data.get('preheader') or '').strip()
    cta_text = (data.get('cta_text') or '').strip() or 'Visit AlumniConnect'

    if not subject:
        return jsonify({'success': False, 'message': 'Subject is required'}), 400
    if not message:
        return jsonify({'success': False, 'message': 'Message is required'}), 400
    if not recipients:
        return jsonify({'success': False, 'message': 'No recipients selected'}), 400

    conn = get_db()
    try:
        result = send_email(recipients, subject, message, preheader, cta_text)
        sent = int(result.get('sent', 0))
        failed = int(result.get('failed', 0))
        status = 'success' if failed == 0 else ('partial' if sent > 0 else 'failed')
        insert_email_log(conn, subject, len(recipients), sent, failed, status, None)
        return jsonify({'success': True, **result})
    except Exception as ex:
        insert_email_log(conn, subject, len(recipients), 0, len(recipients), 'failed', str(ex))
        return jsonify({'success': False, 'message': str(ex)}), 500


@app.route('/api/email/logs', methods=['GET'])
def get_email_logs():
    conn = get_db()
    return jsonify(fetch_email_logs(conn, limit=10))


@app.route('/api/events/<int:eid>', methods=['PUT'])
def update_event(eid):
    data = request.get_json() or {}
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute(
            "UPDATE events SET title=%s, date=%s, event_time=%s, registration_deadline=%s, location=%s, description=%s, fee=%s, payment_account=%s, audience=%s WHERE id=%s",
            (data.get('title'), data.get('date'), data.get('time') or None, data.get('registration_deadline') or None, data.get('location'), data.get('description'),
             data.get('fee', 0) or 0, data.get('payment_account'), data.get('audience', 'both'), eid)
        )
    except pymysql_err.OperationalError as e:
        if e.args and e.args[0] == 1054:  # Unknown column
            try:
                cur.execute(
                    "UPDATE events SET title=%s, date=%s, event_time=%s, location=%s, description=%s, fee=%s, payment_account=%s, audience=%s WHERE id=%s",
                    (data.get('title'), data.get('date'), data.get('time') or None, data.get('location'), data.get('description'),
                     data.get('fee', 0) or 0, data.get('payment_account'), data.get('audience', 'both'), eid)
                )
            except pymysql_err.OperationalError as e2:
                if e2.args and e2.args[0] == 1054:
                    cur.execute(
                        "UPDATE events SET title=%s, date=%s, location=%s, description=%s, fee=%s, payment_account=%s, audience=%s WHERE id=%s",
                        (data.get('title'), data.get('date'), data.get('location'), data.get('description'),
                         data.get('fee', 0) or 0, data.get('payment_account'), data.get('audience', 'both'), eid)
                    )
                else:
                    raise
        else:
            raise
    conn.commit()
    cur.close()
    return jsonify({'success': True})


@app.route('/api/events/<int:eid>/upload-image', methods=['POST', 'OPTIONS'])
@app.route('/api/events/<int:eid>/upload-image/', methods=['POST', 'OPTIONS'])
def upload_event_image(eid):
    # Handle CORS preflight
    if request.method == 'OPTIONS':
        resp = make_response('', 200)
        resp.headers['Access-Control-Allow-Origin'] = '*'
        resp.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return resp

    # Debug: log incoming request shape to help diagnose 404/route-mismatch issues
    try:
        print(f"[UPLOAD] Headers: {dict(request.headers)}")
        print(f"[UPLOAD] Form keys: {list(request.form.keys())}")
        print(f"[UPLOAD] Files keys: {list(request.files.keys())}")
    except Exception:
        pass

    image_file = request.files.get('file') or request.files.get('banner_image')
    if not image_file:
        resp = jsonify({'success': False, 'message': 'Image file is required'})
        resp.status_code = 400
        resp.headers['Access-Control-Allow-Origin'] = '*'
        return resp

    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute('SELECT id FROM events WHERE id=%s', (eid,))
        existing = cur.fetchone()
        if not existing:
            resp = jsonify({'success': False, 'message': 'Event not found'})
            resp.status_code = 404
            resp.headers['Access-Control-Allow-Origin'] = '*'
            return resp

        saved_image = save_uploaded_image(image_file)
        if not saved_image:
            resp = jsonify({'success': False, 'message': 'Invalid image file'})
            resp.status_code = 400
            resp.headers['Access-Control-Allow-Origin'] = '*'
            return resp

        banner_image = build_upload_url(saved_image)

        cur.execute('UPDATE events SET banner_image=%s WHERE id=%s', (banner_image, eid))
        conn.commit()
        resp = jsonify({
            'success': True,
            'banner_image': banner_image,
            'banner_image_url': banner_image
        })
        resp.headers['Access-Control-Allow-Origin'] = '*'
        return resp
    except Exception as ex:
        conn.rollback()
        resp = jsonify({'success': False, 'message': str(ex)})
        resp.status_code = 500
        resp.headers['Access-Control-Allow-Origin'] = '*'
        return resp
    finally:
        cur.close()


@app.route('/events/<int:eid>/upload-image', methods=['POST', 'OPTIONS'])
@app.route('/events/<int:eid>/upload-image/', methods=['POST', 'OPTIONS'])
def upload_event_image_alias(eid):
    """Alias route (no /api prefix) to support clients that POST without the /api base path."""
    return upload_event_image(eid)


@app.route('/api/events/<int:eid>', methods=['DELETE'])
def delete_event(eid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM events WHERE id=%s", (eid,))
    conn.commit()
    cur.close()
    return jsonify({'success': True})


@app.route('/api/events/<int:eid>/register', methods=['POST'])
def register_for_event(eid):
    """Alumni fills a form to register for an event."""
    data = request.get_json()
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute(
            """INSERT INTO event_attendees (event_id, alumni_id, name, student_id, session, email, phone, transaction_id)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
            (eid, data.get('alumni_id'), data.get('name'), data.get('student_id'),
             data.get('session'), data.get('email'), data.get('phone'), data.get('transaction_id'))
        )
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        cur.close()

    return jsonify({'success': True, 'message': 'Registration successful'}), 201


@app.route('/api/alumni/<int:aid>/photo', methods=['POST'])
def update_alumni_photo(aid):
    photo_file = request.files.get('photo')
    if not photo_file:
        return jsonify({'success': False, 'message': 'photo file is required'}), 400

    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("SELECT status FROM alumni WHERE id=%s", (aid,))
        existing = cur.fetchone()
        if not existing:
            return jsonify({'success': False, 'message': 'Alumni not found'}), 404
        if existing.get('status') != 'approved':
            return jsonify({'success': False, 'message': 'Profile photo can be changed after approval'}), 403

        filename = save_uploaded_image(photo_file)
        if not filename:
            return jsonify({'success': False, 'message': 'Invalid image file'}), 400

        cur.execute("UPDATE alumni SET photo=%s WHERE id=%s", (filename, aid))
        conn.commit()
        return jsonify({'success': True, 'photo_url': build_upload_url(filename), 'photo': filename})
    except Exception as ex:
        conn.rollback()
        return jsonify({'success': False, 'message': str(ex)}), 500
    finally:
        cur.close()


@app.route('/api/events/<int:eid>/attendees', methods=['GET'])
def get_event_attendees(eid):
    """Admin gets the list of alumni registered for an event."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """SELECT id, alumni_id, name, student_id, session, email, phone, transaction_id,
                  registered_at
           FROM event_attendees WHERE event_id=%s ORDER BY registered_at DESC""",
        (eid,)
    )
    rows = cur.fetchall()
    cur.close()
    for r in rows:
        if r.get('registered_at'):
            r['registered_at'] = str(r['registered_at'])
    return jsonify(rows)


@app.route('/api/events/<int:eid>/join', methods=['POST'])
def join_event(eid):
    data = request.get_json()
    alumni_id = data.get('alumni_id')
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("INSERT INTO event_registrations (alumni_id, event_id) VALUES (%s,%s)", (alumni_id, eid))
        conn.commit()
    except Exception:
        conn.rollback()
        return jsonify({'success': False, 'message': 'Already joined'}), 409
    finally:
        cur.close()
    return jsonify({'success': True})


@app.route('/api/events/<int:eid>/leave', methods=['POST'])
def leave_event(eid):
    data = request.get_json()
    alumni_id = data.get('alumni_id')
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM event_registrations WHERE alumni_id=%s AND event_id=%s", (alumni_id, eid))
    conn.commit()
    cur.close()
    return jsonify({'success': True})


@app.route('/api/alumni/<int:aid>/events', methods=['GET'])
def alumni_events(aid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT e.* FROM events e
        JOIN event_registrations er ON e.id = er.event_id
        WHERE er.alumni_id = %s
    """, (aid,))
    rows = cur.fetchall()
    cur.close()
    for r in rows:
        if r.get('date'): r['date'] = str(r['date'])
        if r.get('created_at'): r['created_at'] = str(r['created_at'])
        if r.get('banner_image'): r['banner_image_url'] = build_upload_url(r.get('banner_image'))
    return jsonify(rows)


# ═══════════════════════════════════════════════════════
#  FUND TRANSACTIONS
# ═══════════════════════════════════════════════════════

@app.route('/api/transactions', methods=['GET'])
def get_transactions():
    alumni_id = request.args.get('alumni_id', type=int)
    conn = get_db()
    cur = conn.cursor()
    query = """
        SELECT ft.*, fr.title AS request_title
        FROM fund_transactions ft
        LEFT JOIN fund_requests fr ON fr.id = ft.request_id
    """
    params = []
    if alumni_id:
        query += " WHERE ft.alumni_id=%s"
        params.append(alumni_id)
    query += " ORDER BY ft.date DESC, ft.id DESC"
    cur.execute(query, tuple(params))
    rows = cur.fetchall()
    cur.close()
    for r in rows:
        if r.get('date'): r['date'] = str(r['date'])
        if r.get('created_at'): r['created_at'] = str(r['created_at'])
        if r.get('amount'): r['amount'] = float(r['amount'])
    return jsonify(rows)


@app.route('/api/transactions', methods=['POST'])
def add_transaction():
    data = request.get_json() or {}
    donor = (data.get('donor') or '').strip()
    tx_type = (data.get('type') or 'Donation').strip()
    amount = data.get('amount')
    tx_date = data.get('date')
    if not donor or not amount or not tx_date:
        return jsonify({'success': False, 'message': 'donor, amount and date are required'}), 400

    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO fund_transactions
        (donor, type, amount, date, note, request_id, alumni_id, payment_method, payment_reference, created_by_role, status)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """,
        (
            donor,
            tx_type,
            amount,
            tx_date,
            data.get('note'),
            data.get('request_id'),
            data.get('alumni_id'),
            data.get('payment_method'),
            data.get('payment_reference') or data.get('transaction_id'),
            data.get('created_by_role') or 'alumni',
            data.get('status') or 'paid',
        )
    )
    conn.commit()
    new_id = cur.lastrowid
    cur.close()
    return jsonify({'success': True, 'id': new_id}), 201


@app.route('/api/transactions/<int:tid>', methods=['DELETE'])
def delete_transaction(tid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM fund_transactions WHERE id=%s", (tid,))
    conn.commit()
    cur.close()
    return jsonify({'success': True})


@app.route('/api/fund-requests', methods=['GET'])
def get_fund_requests():
    status = (request.args.get('status') or '').strip().lower()
    conn = get_db()
    cur = conn.cursor()
    query = "SELECT * FROM fund_requests"
    params = []
    if status in ('open', 'closed'):
        query += " WHERE status=%s"
        params.append(status)
    query += " ORDER BY id DESC"
    cur.execute(query, tuple(params))
    rows = cur.fetchall()
    cur.close()
    for r in rows:
        if r.get('created_at'):
            r['created_at'] = str(r['created_at'])
        if r.get('target_amount') is not None:
            r['target_amount'] = float(r['target_amount'])
    return jsonify(rows)


@app.route('/api/fund-requests', methods=['POST'])
def add_fund_request():
    data = request.get_json() or {}
    title = (data.get('title') or '').strip()
    purpose = (data.get('purpose') or '').strip()
    target_amount = data.get('target_amount')
    payment_option = (data.get('payment_option') or 'both').strip().lower()
    if not title or not purpose or not target_amount:
        return jsonify({'success': False, 'message': 'title, purpose and target_amount are required'}), 400
    if payment_option not in ('bkash', 'bank', 'both'):
        payment_option = 'both'

    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO fund_requests
        (title, purpose, target_amount, payment_option, bkash_number, bank_account_name, bank_account_number, bank_name, status, created_by)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """,
        (
            title,
            purpose,
            target_amount,
            payment_option,
            data.get('bkash_number'),
            data.get('bank_account_name'),
            data.get('bank_account_number'),
            data.get('bank_name'),
            data.get('status') or 'open',
            data.get('created_by') or 'admin',
        ),
    )
    conn.commit()
    new_id = cur.lastrowid
    cur.close()
    return jsonify({'success': True, 'id': new_id}), 201


@app.route('/api/fund-requests/<int:request_id>', methods=['PUT', 'OPTIONS'])
def update_fund_request(request_id):
    try:
        if request.method == 'OPTIONS':
            return ('', 204)

        data = request.get_json() or {}
        title = (data.get('title') or '').strip()
        purpose = (data.get('purpose') or '').strip()
        target_amount = data.get('target_amount')
        payment_option = (data.get('payment_option') or 'both').strip().lower()
        status = (data.get('status') or 'open').strip().lower()

        if not title or not purpose or not target_amount:
            return jsonify({'success': False, 'message': 'title, purpose and target_amount are required'}), 400
        if payment_option not in ('bkash', 'bank', 'both'):
            payment_option = 'both'
        if status not in ('open', 'closed'):
            status = 'open'

        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            """
            UPDATE fund_requests
            SET title=%s, purpose=%s, target_amount=%s, payment_option=%s,
                bkash_number=%s, bank_account_name=%s, bank_account_number=%s, bank_name=%s, status=%s
            WHERE id=%s
            """,
            (
                title,
                purpose,
                target_amount,
                payment_option,
                data.get('bkash_number'),
                data.get('bank_account_name'),
                data.get('bank_account_number'),
                data.get('bank_name'),
                status,
                request_id,
            ),
        )
        conn.commit()
        updated = cur.rowcount
        cur.close()

        if not updated:
            return jsonify({'success': False, 'message': 'Fund request not found'}), 404
        return jsonify({'success': True, 'id': request_id})
    except Exception as e:
        print(f'[ERROR] update_fund_request: {e}')
        return jsonify({'success': False, 'message': 'Failed to update fund request'}), 500


@app.route('/api/fund-requests/<int:request_id>', methods=['DELETE', 'OPTIONS'])
def delete_fund_request(request_id):
    try:
        if request.method == 'OPTIONS':
            return ('', 204)

        conn = get_db()
        cur = conn.cursor()
        cur.execute('DELETE FROM fund_requests WHERE id=%s', (request_id,))
        conn.commit()
        deleted = cur.rowcount
        cur.close()

        if not deleted:
            return jsonify({'success': False, 'message': 'Fund request not found'}), 404
        return jsonify({'success': True, 'message': 'Fund request deleted'})
    except Exception as e:
        print(f'[ERROR] delete_fund_request: {e}')
        return jsonify({'success': False, 'message': 'Failed to delete fund request'}), 500


# ═══════════════════════════════════════════════════════
#  TRAININGS
# ═══════════════════════════════════════════════════════

@app.route('/api/trainings', methods=['GET'])
def get_trainings():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM trainings ORDER BY date DESC")
    rows = cur.fetchall()
    cur.close()
    for r in rows:
        if r.get('date'): r['date'] = str(r['date'])
        if r.get('created_at'): r['created_at'] = str(r['created_at'])
    return jsonify(rows)


@app.route('/api/trainings', methods=['POST'])
def add_training():
    data = request.get_json()
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO trainings (title, trainer, date, seats, enrolled, status, fee, payment_account, created_by) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
        (data.get('title'), data.get('trainer'), data.get('date'),
         data.get('seats', 30), data.get('enrolled', 0), data.get('status', 'Upcoming'),
         data.get('fee', 0), data.get('payment_account'), data.get('created_by'))
    )
    conn.commit()
    new_id = cur.lastrowid
    cur.close()
    return jsonify({'success': True, 'id': new_id}), 201


@app.route('/api/trainings/<int:tid>', methods=['PUT'])
def update_training(tid):
    data = request.get_json()
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "UPDATE trainings SET title=%s, trainer=%s, date=%s, seats=%s, status=%s, fee=%s, payment_account=%s WHERE id=%s",
        (data.get('title'), data.get('trainer'), data.get('date'),
         data.get('seats'), data.get('status'),
         data.get('fee', 0), data.get('payment_account'), tid)
    )
    conn.commit()
    cur.close()
    return jsonify({'success': True})


@app.route('/api/trainings/<int:tid>', methods=['DELETE'])
def delete_training(tid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM trainings WHERE id=%s", (tid,))
    conn.commit()
    cur.close()
    return jsonify({'success': True})


@app.route('/api/trainings/<int:tid>/register', methods=['POST'])
def register_for_training(tid):
    data = request.get_json()
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT seats, enrolled FROM trainings WHERE id=%s", (tid,))
    t = cur.fetchone()
    if not t:
        cur.close()
        return jsonify({'success': False, 'message': 'Training not found'}), 404
    if t['enrolled'] >= t['seats']:
        cur.close()
        return jsonify({'success': False, 'message': 'Training is full'}), 400
    try:
        cur.execute(
            "INSERT INTO training_attendees (training_id, alumni_id, name, student_id, email, phone, payment_method, transaction_id) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
            (tid, data.get('alumni_id'), data.get('name'), data.get('student_id'),
             data.get('email'), data.get('phone'), data.get('payment_method'), data.get('transaction_id'))
        )
        cur.execute("UPDATE trainings SET enrolled = enrolled + 1 WHERE id=%s", (tid,))
        conn.commit()
    except Exception as ex:
        conn.rollback()
        cur.close()
        return jsonify({'success': False, 'message': str(ex)}), 409
    cur.close()
    return jsonify({'success': True}), 201


@app.route('/api/trainings/<int:tid>/attendees', methods=['GET'])
def get_training_attendees(tid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM training_attendees WHERE training_id=%s ORDER BY registered_at DESC", (tid,))
    rows = cur.fetchall()
    cur.close()
    for r in rows:
        if r.get('registered_at'): r['registered_at'] = str(r['registered_at'])
    return jsonify(rows)


@app.route('/api/alumni/<int:aid>/enrolled-training-ids', methods=['GET'])
def get_enrolled_training_ids(aid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT training_id FROM training_attendees WHERE alumni_id=%s", (aid,))
    rows = cur.fetchall()
    cur.close()
    return jsonify([r['training_id'] for r in rows])


@app.route('/api/alumni/<int:aid>/registered-event-ids', methods=['GET'])
def get_registered_event_ids(aid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT event_id FROM event_attendees WHERE alumni_id=%s", (aid,))
    rows = cur.fetchall()
    cur.close()
    return jsonify([r['event_id'] for r in rows])


@app.route('/api/trainings/<int:tid>/enroll', methods=['POST'])
def enroll_training(tid):
    data = request.get_json()
    alumni_id = data.get('alumni_id')
    conn = get_db()
    cur = conn.cursor()
    # check seats
    cur.execute("SELECT seats, enrolled FROM trainings WHERE id=%s", (tid,))
    t = cur.fetchone()
    if not t:
        cur.close()
        return jsonify({'success': False, 'message': 'Training not found'}), 404
    if t['enrolled'] >= t['seats']:
        cur.close()
        return jsonify({'success': False, 'message': 'Training is full'}), 400
    try:
        cur.execute("INSERT INTO training_enrollments (alumni_id, training_id) VALUES (%s,%s)", (alumni_id, tid))
        cur.execute("UPDATE trainings SET enrolled = enrolled + 1 WHERE id=%s", (tid,))
        conn.commit()
    except Exception:
        conn.rollback()
        return jsonify({'success': False, 'message': 'Already enrolled'}), 409
    finally:
        cur.close()
    return jsonify({'success': True})


@app.route('/api/trainings/<int:tid>/unenroll', methods=['POST'])
def unenroll_training(tid):
    data = request.get_json()
    alumni_id = data.get('alumni_id')
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM training_enrollments WHERE alumni_id=%s AND training_id=%s", (alumni_id, tid))
    if cur.rowcount > 0:
        cur.execute("UPDATE trainings SET enrolled = GREATEST(0, enrolled - 1) WHERE id=%s", (tid,))
    conn.commit()
    cur.close()
    return jsonify({'success': True})


@app.route('/api/alumni/<int:aid>/trainings', methods=['GET'])
def alumni_trainings(aid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT t.* FROM trainings t
        JOIN training_enrollments te ON t.id = te.training_id
        WHERE te.alumni_id = %s
    """, (aid,))
    rows = cur.fetchall()
    cur.close()
    for r in rows:
        if r.get('date'): r['date'] = str(r['date'])
        if r.get('created_at'): r['created_at'] = str(r['created_at'])
    return jsonify(rows)


# ═══════════════════════════════════════════════════════
#  JOBS
# ═══════════════════════════════════════════════════════

@app.route('/api/jobs', methods=['GET'])
def get_jobs():
    """Returns only approved jobs (used by alumni dashboard and admin listing)."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM jobs WHERE status='approved' ORDER BY created_at DESC")
    rows = cur.fetchall()
    cur.close()
    for r in rows:
        if r.get('deadline'):   r['deadline']   = str(r['deadline'])
        if r.get('created_at'): r['created_at'] = str(r['created_at'])
    return jsonify(rows)


@app.route('/api/jobs/pending-submissions', methods=['GET'])
def get_pending_jobs():
    """Returns alumni-submitted jobs awaiting admin approval."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT j.*, a.name AS alumni_name, a.email AS alumni_email
        FROM jobs j
        LEFT JOIN alumni a ON a.id = j.submitted_by
        WHERE j.status='pending'
        ORDER BY j.created_at DESC
    """)
    rows = cur.fetchall()
    cur.close()
    for r in rows:
        if r.get('deadline'):   r['deadline']   = str(r['deadline'])
        if r.get('created_at'): r['created_at'] = str(r['created_at'])
    return jsonify(rows)


@app.route('/api/jobs', methods=['POST'])
def add_job():
    """Admin posts a job – immediately approved."""
    d = request.get_json()
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        'INSERT INTO jobs (title, company, location, type, deadline, description, apply_link, status) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)',
        (d['title'], d.get('company',''), d.get('location',''), d.get('type','Full-time'),
         d.get('deadline') or None, d.get('description',''), d.get('apply_link',''), 'approved')
    )
    conn.commit()
    new_id = cur.lastrowid
    cur.close()
    return jsonify({'success': True, 'id': new_id}), 201


@app.route('/api/jobs/submit', methods=['POST'])
def submit_job():
    """Alumni submits a job for admin approval."""
    d = request.get_json()
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        'INSERT INTO jobs (title, company, location, type, deadline, description, apply_link, submitted_by, status) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)',
        (d['title'], d.get('company',''), d.get('location',''), d.get('type','Full-time'),
         d.get('deadline') or None, d.get('description',''), d.get('apply_link',''),
         d.get('submitted_by') or None, 'pending')
    )
    conn.commit()
    new_id = cur.lastrowid
    cur.close()
    return jsonify({'success': True, 'id': new_id}), 201


@app.route('/api/jobs/<int:jid>/approve', methods=['POST'])
def approve_job(jid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE jobs SET status='approved' WHERE id=%s", (jid,))
    conn.commit()
    cur.close()
    return jsonify({'success': True})


@app.route('/api/jobs/<int:jid>', methods=['DELETE'])
def delete_job(jid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute('DELETE FROM jobs WHERE id=%s', (jid,))
    conn.commit()
    cur.close()
    return jsonify({'success': True})


# ═══════════════════════════════════════════════════════
#  EXISTING LISTS (ADMIN EXCEL UPLOADS)
# ═══════════════════════════════════════════════════════

@app.route('/api/existing-lists', methods=['GET'])
def get_existing_lists():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, title, file_name, stored_path, uploaded_by, created_at FROM existing_lists ORDER BY created_at DESC, id DESC")
    rows = cur.fetchall()
    cur.close()

    for r in rows:
        r['file_url'] = build_upload_url(r.get('stored_path'))
        if r.get('created_at'):
            r['created_at'] = str(r['created_at'])

    return jsonify(rows)


@app.route('/api/existing-lists', methods=['POST'])
def upload_existing_list():
    content_type = (request.content_type or '').lower()
    is_multipart = ('multipart/form-data' in content_type) or bool(request.files)
    if not is_multipart:
        return jsonify({'success': False, 'message': 'Upload must be multipart/form-data'}), 400

    excel_file = request.files.get('file')
    if not excel_file or not excel_file.filename:
        return jsonify({'success': False, 'message': 'Excel file is required'}), 400

    original_name = secure_filename(excel_file.filename)
    if '.' not in original_name:
        return jsonify({'success': False, 'message': 'Invalid file name'}), 400

    ext = original_name.rsplit('.', 1)[-1].lower()
    if ext not in ALLOWED_EXISTING_LIST_EXTENSIONS:
        return jsonify({'success': False, 'message': 'Only .xlsx or .xls files are allowed'}), 400

    list_title = os.path.splitext(original_name)[0].strip() or 'Untitled List'
    stored_name = f"{uuid.uuid4().hex}_{original_name}"
    subdir = os.path.join(UPLOAD_FOLDER, 'existing_lists')
    os.makedirs(subdir, exist_ok=True)
    full_path = os.path.join(subdir, stored_name)
    excel_file.save(full_path)

    stored_path = f"existing_lists/{stored_name}"

    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO existing_lists (title, file_name, stored_path, uploaded_by) VALUES (%s,%s,%s,%s)",
        (list_title, original_name, stored_path, 'admin')
    )
    conn.commit()
    new_id = cur.lastrowid
    cur.close()

    return jsonify({
        'success': True,
        'id': new_id,
        'title': list_title,
        'file_name': original_name,
        'file_url': build_upload_url(stored_path),
    }), 201


@app.route('/api/existing-lists/<int:list_id>', methods=['DELETE'])
def delete_existing_list(list_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT stored_path FROM existing_lists WHERE id=%s", (list_id,))
    row = cur.fetchone()
    if not row:
        cur.close()
        return jsonify({'success': False, 'message': 'List not found'}), 404

    cur.execute("DELETE FROM existing_lists WHERE id=%s", (list_id,))
    conn.commit()
    cur.close()

    stored_path = row.get('stored_path')
    if stored_path:
        file_path = os.path.join(UPLOAD_FOLDER, stored_path)
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
        except Exception:
            pass

    return jsonify({'success': True})


@app.route('/api/existing-lists/<int:list_id>/data', methods=['GET'])
def get_existing_list_data(list_id):
    """Read and return the data from an Excel file"""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT stored_path, title FROM existing_lists WHERE id=%s", (list_id,))
    row = cur.fetchone()
    cur.close()

    if not row:
        return jsonify({'success': False, 'message': 'List not found'}), 404

    stored_path = row.get('stored_path')
    file_path = os.path.join(UPLOAD_FOLDER, stored_path)

    if not os.path.exists(file_path):
        return jsonify({'success': False, 'message': 'File not found on disk'}), 404

    try:
        # Load Excel workbook
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        rows = []
        headers = []

        # Extract rows and headers
        for idx, row_cells in enumerate(ws.iter_rows(values_only=True), 1):
            if idx == 1:
                # First row is headers
                headers = [str(cell).strip() if cell else f'Column {i+1}' for i, cell in enumerate(row_cells)]
            else:
                # Data rows
                row_data = {}
                for i, cell in enumerate(row_cells):
                    col_name = headers[i] if i < len(headers) else f'Column {i+1}'
                    row_data[col_name] = cell if cell is not None else ''
                rows.append(row_data)

        wb.close()

        return jsonify({
            'success': True,
            'title': row.get('title'),
            'headers': headers,
            'rows': rows,
            'total': len(rows)
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error reading Excel file: {str(e)}'}), 400


# ═══════════════════════════════════════════════════════
#  EXISTING ALUMNI (ADMIN-ADDED ALUMNI LIST)
# ═══════════════════════════════════════════════════════

@app.route('/api/exist-alumni', methods=['GET'])
def get_exist_alumni():
    """Get all manually-added alumni by admin"""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, name, email, phone, student_id, session, department, created_at
        FROM alumni
        WHERE status='approved' AND is_manually_added=1
        ORDER BY created_at DESC
    """)
    rows = cur.fetchall()
    cur.close()
    for r in rows:
        if r.get('created_at'):
            r['created_at'] = str(r['created_at'])
    return jsonify(rows)


@app.route('/api/exist-alumni', methods=['POST'])
def add_exist_alumni():
    """Add a single alumni manually (name, student_id, session, email, phone)"""
    data = request.get_json() or {}
    
    required_fields = ['name', 'student_id', 'email', 'session', 'phone']
    for field in required_fields:
        if not data.get(field) or not str(data.get(field, '')).strip():
            return jsonify({'success': False, 'message': f'{field} is required'}), 400

    name = str(data['name']).strip()
    student_id = str(data['student_id']).strip()
    email = str(data['email']).strip()
    session = str(data['session']).strip()
    phone = str(data['phone']).strip()
    department = str(data.get('department', 'ICE')).strip() or 'ICE'

    conn = get_db()
    cur = conn.cursor()
    try:
        # Use a dummy password since admin adds them directly
        dummy_password = generate_password_hash('ADMIN_ADDED_' + str(uuid.uuid4())[:8])
        
        cur.execute(
            """INSERT INTO alumni (name, email, phone, student_id, session, department, 
                                  password, status, user_type, is_manually_added)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
            (name, email, phone, student_id, session, department, dummy_password, 'approved', 'alumni', 1)
        )
        conn.commit()
        new_id = cur.lastrowid
        cur.close()
        
        return jsonify({
            'success': True,
            'id': new_id,
            'message': f'Alumni "{name}" added successfully'
        }), 201
    
    except Exception as e:
        conn.rollback()
        cur.close()
        if 'Duplicate entry' in str(e) and 'email' in str(e):
            return jsonify({'success': False, 'message': 'Email already exists'}), 409
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/exist-alumni/bulk', methods=['POST'])
def bulk_add_exist_alumni():
    """Bulk add alumni from Excel file data"""
    data = request.get_json() or {}
    alumni_records = data.get('alumni_records') or []
    
    if not alumni_records or not isinstance(alumni_records, list):
        return jsonify({'success': False, 'message': 'alumni_records must be a non-empty list'}), 400
    
    conn = get_db()
    cur = conn.cursor()
    added_count = 0
    errors = []
    
    for idx, record in enumerate(alumni_records):
        try:
            name = str(record.get('name', '')).strip()
            student_id = str(record.get('student_id', '') or record.get('id', '')).strip()
            email = str(record.get('email', '')).strip()
            session = str(record.get('session', '')).strip()
            phone = str(record.get('phone', '')).strip()
            department = str(record.get('department', 'ICE')).strip() or 'ICE'
            
            # Validate required fields
            if not name or not student_id or not email or not session or not phone:
                errors.append({
                    'row': idx + 1,
                    'reason': 'Missing required fields (name, student_id, email, session, phone)'
                })
                continue
            
            dummy_password = generate_password_hash('ADMIN_ADDED_' + str(uuid.uuid4())[:8])
            
            cur.execute(
                """INSERT INTO alumni (name, email, phone, student_id, session, department,
                                      password, status, user_type, is_manually_added)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (name, email, phone, student_id, session, department, dummy_password, 'approved', 'alumni', 1)
            )
            added_count += 1
        
        except Exception as e:
            error_msg = str(e)
            if 'Duplicate entry' in error_msg and 'email' in error_msg:
                error_msg = f'Email "{record.get("email")}" already exists'
            errors.append({
                'row': idx + 1,
                'name': record.get('name', 'Unknown'),
                'reason': error_msg
            })
    
    try:
        conn.commit()
    except Exception as e:
        conn.rollback()
        cur.close()
        return jsonify({
            'success': False,
            'message': f'Error committing records: {str(e)}',
            'added': added_count,
            'errors': errors
        }), 500
    
    cur.close()
    
    return jsonify({
        'success': True,
        'added': added_count,
        'total_attempted': len(alumni_records),
        'errors': errors if errors else [],
        'message': f'Successfully added {added_count} alumni out of {len(alumni_records)} records'
    }), 201


@app.route('/api/exist-alumni/<int:alumni_id>', methods=['DELETE'])
def delete_exist_alumni(alumni_id):
    """Delete an existing alumni record"""
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("DELETE FROM alumni WHERE id=%s", (alumni_id,))
    conn.commit()
    cur.close()
    
    return jsonify({'success': True, 'message': 'Alumni record deleted successfully'})



# ═══════════════════════════════════════════════════════
#  STATS (dashboard summary)
# ═══════════════════════════════════════════════════════

@app.route('/api/stats', methods=['GET'])
def get_stats():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) AS cnt FROM alumni WHERE status='approved' AND (user_type IS NULL OR user_type='alumni') AND (is_manually_added IS NULL OR is_manually_added=0)")
    total_alumni = cur.fetchone()['cnt']
    cur.execute("SELECT COUNT(*) AS cnt FROM alumni WHERE status='approved' AND user_type='student'")
    total_students = cur.fetchone()['cnt']
    cur.execute("SELECT COUNT(*) AS cnt FROM alumni WHERE status='pending'")
    pending_cnt = cur.fetchone()['cnt']
    cur.execute("SELECT COUNT(*) AS cnt FROM events")
    events_cnt = cur.fetchone()['cnt']
    cur.execute("SELECT COALESCE(SUM(amount),0) AS total FROM fund_transactions")
    total_funds = float(cur.fetchone()['total'])
    cur.execute("SELECT COUNT(*) AS cnt FROM jobs")
    jobs_cnt = cur.fetchone()['cnt']
    cur.close()
    return jsonify({
        'total_alumni':   total_alumni,
        'total_students': total_students,
        'pending':        pending_cnt,
        'events':         events_cnt,
        'total_funds':    total_funds,
        'total_jobs':     jobs_cnt,
    })


@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'success': True, 'message': 'ok'})


@app.route('/api/version', methods=['GET'])
def version():
    """Diagnostic endpoint to check if latest code is deployed and DB status."""
    db_status = 'unknown'
    db_version = None
    
    try:
        if _db_ready:
            db_status = 'ready'
        elif _db_init_attempted:
            db_status = 'init_attempted_failed'
        else:
            db_status = 'not_attempted'
    except Exception as e:
        db_status = f'error: {str(e)[:50]}'
    
    return jsonify({
        'success': True,
        'code_version': '1b515e4',  # Updated after DB retry loop fix
        'db_init_attempted': _db_init_attempted,
        'db_ready': _db_ready,
        'db_status': db_status,
    })


@app.route('/', methods=['GET'])
def root():
    return jsonify({
        'success': True,
        'message': 'AlumniConnect API is running',
        'health': '/api/health'
    })


@app.route('/api', methods=['GET'])
def api_root():
    return jsonify({
        'success': True,
        'message': 'AlumniConnect API root',
        'health': '/api/health'
    })


# ─────────────────────────────────────────────────────
# Success Stories API
# ─────────────────────────────────────────────────────

@app.route('/api/success-stories', methods=['GET'])
def get_success_stories():
    """Get paginated success stories with alumni info"""
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 10))
        offset = (page - 1) * limit

        conn = get_db()
        cur = conn.cursor()

        # Fetch success stories with alumni info
        cur.execute("""
            SELECT 
                ss.id,
                ss.alumni_id,
                ss.title,
                ss.story,
                ss.current_position,
                ss.batch,
                ss.department,
                ss.image_url,
                ss.created_at,
                a.id as alumni_id,
                a.name as alumni_name,
                a.photo as alumni_photo
            FROM success_stories ss
            LEFT JOIN alumni a ON ss.alumni_id = a.id
            WHERE a.status = 'approved'
            ORDER BY ss.created_at DESC
            LIMIT %s OFFSET %s
        """, (limit, offset))

        rows = cur.fetchall()
        cur.close()

        stories = []
        for row in rows:
            story = {
                'id': row.get('id'),
                'alumni_id': row.get('alumni_id'),
                'title': row.get('title'),
                'story': row.get('story'),
                'current_position': row.get('current_position'),
                'batch': row.get('batch'),
                'department': row.get('department'),
                'image_url': row.get('image_url'),
                'created_at': str(row.get('created_at')) if row.get('created_at') else None,
                'alumni': {
                    'id': row.get('alumni_id'),
                    'name': row.get('alumni_name'),
                    'photo': row.get('alumni_photo'),
                }
            }
            stories.append(story)

        return jsonify({
            'success': True,
            'data': stories,
            'page': page,
            'limit': limit
        })

    except Exception as e:
        print(f'[ERROR] get_success_stories: {e}')
        return jsonify({
            'success': False,
            'message': 'Failed to fetch success stories'
        }), 500


@app.route('/api/success-stories', methods=['POST'])
def create_success_story():
    """Create a new success story with optional image upload"""
    try:
        # Get alumni_id or student_id from form data
        alumni_id_str = request.form.get('alumni_id', '').strip()
        student_id_str = request.form.get('student_id', '').strip()

        print(f'[DEBUG POST] alumni_id_str={repr(alumni_id_str)}, student_id_str={repr(student_id_str)}')

        # Determine which identifier was provided
        alumni_id = None
        if alumni_id_str:
            try:
                alumni_id = int(alumni_id_str)
            except (ValueError, TypeError):
                pass

        # If student_id was provided, look up the alumni_id
        if not alumni_id and student_id_str:
            try:
                conn = get_db()
                cur = conn.cursor()
                print(f'[DEBUG] Executing query with student_id={repr(student_id_str)}')
                cur.execute('SELECT id FROM alumni WHERE student_id = %s', (student_id_str,))
                result = cur.fetchone()
                print(f'[DEBUG] Query result: {result}')
                if result:
                    alumni_id = result['id'] if isinstance(result, dict) else result[0]
                print(f'[DEBUG] Looked up student_id={repr(student_id_str)}, found alumni_id={alumni_id}')
            except Exception as lookup_err:
                import traceback
                print(f'[WARNING] Failed to lookup alumni by student_id: {lookup_err}')
                print(f'[TRACEBACK]\n{traceback.format_exc()}')

        if not alumni_id:
            return jsonify({
                'success': False,
                'message': 'Alumni ID or Student ID is required'
            }), 400

        # Parse form data
        title = request.form.get('title', '').strip() or None
        story = request.form.get('story', '').strip()
        current_position = request.form.get('current_position', '').strip() or None
        batch = request.form.get('batch', '').strip()
        department = request.form.get('department', 'ICE').strip()
        image_file = request.files.get('image')

        # Validate required fields
        if not story:
            return jsonify({
                'success': False,
                'message': 'Story text is required'
            }), 400

        if not batch:
            return jsonify({
                'success': False,
                'message': 'Batch is required'
            }), 400

        # Upload image if provided
        image_url = None
        if image_file and image_file.filename:
            if allowed_file(image_file.filename, {'png', 'jpg', 'jpeg', 'gif', 'webp'}):
                try:
                    upload_result = upload_to_cloudinary(image_file.stream, image_file.filename, 'success_stories')
                    # upload_to_cloudinary returns a dict on success or error info on failure
                    if isinstance(upload_result, dict):
                        if upload_result.get('error'):
                            print(f'[WARNING] Image upload error: {upload_result.get("error")}')
                        else:
                            image_url = upload_result.get('url') or upload_result.get('secure_url')
                    elif isinstance(upload_result, str):
                        image_url = upload_result
                except Exception as e:
                    print(f'[WARNING] Image upload failed: {e}')
                    # Continue without image

        # Insert into database
        conn = get_db()
        cur = conn.cursor()

        # Debug: log param types before executing insert
        try:
            print('[DEBUG INSERT] params types:', type(alumni_id), type(title), type(story), type(current_position), type(batch), type(department), type(image_url))
        except Exception:
            pass

        cur.execute("""
            INSERT INTO success_stories (
                alumni_id, title, story, current_position, batch, department, image_url, created_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, NOW())
        """, (
            alumni_id, title, story, current_position, batch, department, image_url
        ))

        conn.commit()
        new_id = cur.lastrowid
        cur.close()

        return jsonify({
            'success': True,
            'message': 'Success story created',
            'story_id': new_id
        }), 201

    except Exception as e:
        import traceback
        error_msg = f'{type(e).__name__}: {str(e)}'
        print(f'[ERROR] create_success_story: {error_msg}')
        print(f'[TRACEBACK]\n{traceback.format_exc()}')
        return jsonify({
            'success': False,
            'message': 'Failed to create success story'
        }), 500


@app.route('/api/success-stories/mine', methods=['GET'])
def get_my_success_stories():
    """Get success stories posted by a specific alumni (owner).

    Accepts either `alumni_id` or `student_id` as query params. Returns
    stories belonging to that alumni ordered by created_at desc.
    """
    try:
        alumni_id = request.args.get('alumni_id', '').strip()
        student_id = request.args.get('student_id', '').strip()

        resolved_alumni_id = None
        if alumni_id:
            try:
                resolved_alumni_id = int(alumni_id)
            except Exception:
                resolved_alumni_id = None

        if not resolved_alumni_id and student_id:
            conn = get_db()
            cur = conn.cursor()
            cur.execute('SELECT id FROM alumni WHERE student_id=%s LIMIT 1', (student_id,))
            r = cur.fetchone()
            cur.close()
            if r:
                resolved_alumni_id = r['id'] if isinstance(r, dict) else r[0]

        if not resolved_alumni_id:
            return jsonify({'success': False, 'message': 'alumni_id or student_id required'}), 400

        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT ss.id, ss.alumni_id, ss.title, ss.story, ss.current_position, ss.batch, ss.department, ss.image_url, ss.created_at,
                   a.name as alumni_name, a.photo as alumni_photo
            FROM success_stories ss
            LEFT JOIN alumni a ON ss.alumni_id = a.id
            WHERE ss.alumni_id = %s
            ORDER BY ss.created_at DESC
        """, (resolved_alumni_id,))
        rows = cur.fetchall()
        cur.close()

        stories = []
        for row in rows:
            stories.append({
                'id': row.get('id'),
                'alumni_id': row.get('alumni_id'),
                'title': row.get('title'),
                'story': row.get('story'),
                'current_position': row.get('current_position'),
                'batch': row.get('batch'),
                'department': row.get('department'),
                'image_url': row.get('image_url'),
                'created_at': str(row.get('created_at')) if row.get('created_at') else None,
                'alumni': {
                    'id': row.get('alumni_id'),
                    'name': row.get('alumni_name'),
                    'photo': row.get('alumni_photo'),
                }
            })

        return jsonify({'success': True, 'data': stories})
    except Exception as e:
        print(f'[ERROR] get_my_success_stories: {e}')
        return jsonify({'success': False, 'message': 'Failed to fetch user stories'}), 500



@app.route('/api/success-stories/<int:story_id>', methods=['PUT', 'OPTIONS'])
def update_success_story(story_id):
    """Update a success story. Ownership must be verified via alumni_id or student_id passed in the request (form or JSON).
    """
    try:
        if request.method == 'OPTIONS':
            return ('', 204)

        # Accept JSON or form data
        data = request.get_json(silent=True) or request.form or {}
        alumni_id_str = data.get('alumni_id') or request.args.get('alumni_id')
        student_id_str = data.get('student_id') or request.args.get('student_id')

        resolved_alumni_id = None
        if alumni_id_str:
            try:
                resolved_alumni_id = int(alumni_id_str)
            except Exception:
                resolved_alumni_id = None

        if not resolved_alumni_id and student_id_str:
            conn = get_db()
            cur = conn.cursor()
            cur.execute('SELECT id FROM alumni WHERE student_id=%s LIMIT 1', (student_id_str,))
            r = cur.fetchone()
            cur.close()
            if r:
                resolved_alumni_id = r['id'] if isinstance(r, dict) else r[0]

        if not resolved_alumni_id:
            return jsonify({'success': False, 'message': 'alumni_id or student_id required for ownership verification'}), 400

        # Verify ownership
        conn = get_db()
        cur = conn.cursor()
        cur.execute('SELECT alumni_id FROM success_stories WHERE id=%s LIMIT 1', (story_id,))
        row = cur.fetchone()
        if not row:
            cur.close()
            return jsonify({'success': False, 'message': 'Story not found'}), 404
        owner_id = row['alumni_id'] if isinstance(row, dict) else row[0]
        if int(owner_id) != int(resolved_alumni_id):
            cur.close()
            return jsonify({'success': False, 'message': 'Not authorized to edit this story'}), 403

        # Perform update (only allow certain fields)
        title = (data.get('title') or request.form.get('title') or '').strip() or None
        story_text = (data.get('story') or request.form.get('story') or '').strip() or None
        current_position = (data.get('current_position') or request.form.get('current_position') or '').strip() or None
        batch = (data.get('batch') or request.form.get('batch') or '').strip() or None
        department = (data.get('department') or request.form.get('department') or '').strip() or None

        # Build dynamic update
        updates = []
        params = []
        if title is not None:
            updates.append('title=%s'); params.append(title)
        if story_text is not None:
            updates.append('story=%s'); params.append(story_text)
        if current_position is not None:
            updates.append('current_position=%s'); params.append(current_position)
        if batch is not None:
            updates.append('batch=%s'); params.append(batch)
        if department is not None:
            updates.append('department=%s'); params.append(department)

        if not updates:
            cur.close()
            return jsonify({'success': False, 'message': 'No updatable fields provided'}), 400

        sql = f"UPDATE success_stories SET {', '.join(updates)} WHERE id=%s"
        params.append(story_id)
        cur.execute(sql, tuple(params))
        conn.commit()
        cur.close()

        return jsonify({'success': True, 'message': 'Story updated'})
    except Exception as e:
        print(f'[ERROR] update_success_story: {e}')
        return jsonify({'success': False, 'message': 'Failed to update story'}), 500


@app.route('/api/success-stories/<int:story_id>', methods=['DELETE', 'OPTIONS'])
def delete_success_story(story_id):
    """Delete a success story. Ownership must be verified via alumni_id or student_id passed in the request (query param or JSON/form).
    """
    try:
        if request.method == 'OPTIONS':
            return ('', 204)

        alumni_id = request.args.get('alumni_id') or request.form.get('alumni_id') or (request.get_json(silent=True) or {}).get('alumni_id')
        student_id = request.args.get('student_id') or request.form.get('student_id') or (request.get_json(silent=True) or {}).get('student_id')

        resolved_alumni_id = None
        if alumni_id:
            try:
                resolved_alumni_id = int(alumni_id)
            except Exception:
                resolved_alumni_id = None

        if not resolved_alumni_id and student_id:
            conn = get_db()
            cur = conn.cursor()
            cur.execute('SELECT id FROM alumni WHERE student_id=%s LIMIT 1', (student_id,))
            r = cur.fetchone()
            cur.close()
            if r:
                resolved_alumni_id = r['id'] if isinstance(r, dict) else r[0]

        if not resolved_alumni_id:
            return jsonify({'success': False, 'message': 'alumni_id or student_id required for ownership verification'}), 400

        conn = get_db()
        cur = conn.cursor()
        cur.execute('SELECT alumni_id FROM success_stories WHERE id=%s LIMIT 1', (story_id,))
        row = cur.fetchone()
        if not row:
            cur.close()
            return jsonify({'success': False, 'message': 'Story not found'}), 404
        owner_id = row['alumni_id'] if isinstance(row, dict) else row[0]
        if int(owner_id) != int(resolved_alumni_id):
            cur.close()
            return jsonify({'success': False, 'message': 'Not authorized to delete this story'}), 403

        cur.execute('DELETE FROM success_stories WHERE id=%s', (story_id,))
        conn.commit()
        cur.close()

        return jsonify({'success': True, 'message': 'Story deleted'})
    except Exception as e:
        print(f'[ERROR] delete_success_story: {e}')
        return jsonify({'success': False, 'message': 'Failed to delete story'}), 500


if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=config.DEBUG, port=config.PORT, use_reloader=False)
 
